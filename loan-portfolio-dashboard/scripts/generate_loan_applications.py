"""
Loan Application Synthetic Data Generator
Generates realistic Indian loan application records and a formula-driven Excel summary.
Usage: python generate_loan_applications.py [--records 3000] [--output /path/to/folder]
"""

import argparse
import os
import random
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--records",        type=int, default=3000,
                    help="Number of application records (default: 3000)")
parser.add_argument("--output",         default=".",
                    help="Output directory")
parser.add_argument("--product-weights", default="40,25,20,15",
                    help="Comma-separated weights: Personal,Home,Auto,Business (default: 40,25,20,15)")
parser.add_argument("--emp-weights",     default="50,30,20",
                    help="Comma-separated weights: Salaried,Self-Employed,Business Owner (default: 50,30,20)")
parser.add_argument("--tier-weights",    default="",
                    help="Comma-separated %% targets: Tier1,Tier2,Tier3 e.g. 30,40,30 "
                         "(overrides per-city weights; must sum to 100)")
parser.add_argument("--loan-amt-personal", default="50000,1500000",
                    help="Min,Max loan amount for Personal loans in Rs (default: 50000,1500000)")
parser.add_argument("--loan-amt-home",     default="1000000,10000000",
                    help="Min,Max for Home loans (default: 1000000,10000000)")
parser.add_argument("--loan-amt-auto",     default="300000,2000000",
                    help="Min,Max for Auto loans (default: 300000,2000000)")
parser.add_argument("--loan-amt-business", default="500000,5000000",
                    help="Min,Max for Business loans (default: 500000,5000000)")
args = parser.parse_args()

NUM     = args.records
OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)

np.random.seed(42)
random.seed(42)

def _parse_weights(s):
    return [float(x.strip()) for x in s.split(",")]

def _parse_range(s):
    parts = [int(x.strip()) for x in s.split(",")]
    return parts[0], parts[1]

PRODUCT_WEIGHTS = _parse_weights(args.product_weights)
EMP_WEIGHTS     = _parse_weights(args.emp_weights)
AMT_PERSONAL    = _parse_range(args.loan_amt_personal)
AMT_HOME        = _parse_range(args.loan_amt_home)
AMT_AUTO        = _parse_range(args.loan_amt_auto)
AMT_BUSINESS    = _parse_range(args.loan_amt_business)

# Print active distribution so it's visible in run output
print(f"  Portfolio distribution applied:")
print(f"    Products   : Personal {PRODUCT_WEIGHTS[0]:.0f}%  Home {PRODUCT_WEIGHTS[1]:.0f}%  "
      f"Auto {PRODUCT_WEIGHTS[2]:.0f}%  Business {PRODUCT_WEIGHTS[3]:.0f}%")
print(f"    Employment : Salaried {EMP_WEIGHTS[0]:.0f}%  SEP {EMP_WEIGHTS[1]:.0f}%  "
      f"Business Owner {EMP_WEIGHTS[2]:.0f}%")
print(f"    Records    : {NUM:,}")
print()

# ── REFERENCE DATA ────────────────────────────────────────────────────────────
CITY_DATA = [
    ("Mumbai",          "Tier 1", "Maharashtra"),
    ("Delhi",           "Tier 1", "Delhi"),
    ("Bengaluru",       "Tier 1", "Karnataka"),
    ("Chennai",         "Tier 1", "Tamil Nadu"),
    ("Hyderabad",       "Tier 1", "Telangana"),
    ("Kolkata",         "Tier 1", "West Bengal"),
    ("Pune",            "Tier 1", "Maharashtra"),
    ("Ahmedabad",       "Tier 1", "Gujarat"),
    ("Jaipur",          "Tier 2", "Rajasthan"),
    ("Lucknow",         "Tier 2", "Uttar Pradesh"),
    ("Kanpur",          "Tier 2", "Uttar Pradesh"),
    ("Nagpur",          "Tier 2", "Maharashtra"),
    ("Indore",          "Tier 2", "Madhya Pradesh"),
    ("Bhopal",          "Tier 2", "Madhya Pradesh"),
    ("Patna",           "Tier 2", "Bihar"),
    ("Vadodara",        "Tier 2", "Gujarat"),
    ("Surat",           "Tier 2", "Gujarat"),
    ("Coimbatore",      "Tier 2", "Tamil Nadu"),
    ("Kochi",           "Tier 2", "Kerala"),
    ("Visakhapatnam",   "Tier 2", "Andhra Pradesh"),
    ("Nashik",          "Tier 2", "Maharashtra"),
    ("Ludhiana",        "Tier 2", "Punjab"),
    ("Madurai",         "Tier 3", "Tamil Nadu"),
    ("Varanasi",        "Tier 3", "Uttar Pradesh"),
    ("Agra",            "Tier 3", "Uttar Pradesh"),
    ("Meerut",          "Tier 3", "Uttar Pradesh"),
    ("Rajkot",          "Tier 3", "Gujarat"),
    ("Jodhpur",         "Tier 3", "Rajasthan"),
    ("Tiruchirappalli", "Tier 3", "Tamil Nadu"),
    ("Mysuru",          "Tier 3", "Karnataka"),
    ("Guwahati",        "Tier 3", "Assam"),
    ("Ranchi",          "Tier 3", "Jharkhand"),
    ("Raipur",          "Tier 3", "Chhattisgarh"),
    ("Dehradun",        "Tier 3", "Uttarakhand"),
    ("Amritsar",        "Tier 3", "Punjab"),
]
if args.tier_weights:
    # Convert user-supplied tier % targets → per-city weights
    tw = _parse_weights(args.tier_weights)   # [T1%, T2%, T3%]
    w1 = tw[0] / 8;  w2 = tw[1] / 14;  w3 = tw[2] / 13
    CITY_WEIGHTS = [w1] * 8 + [w2] * 14 + [w3] * 13
    print(f"    City Tiers : Tier 1 {tw[0]:.0f}%  Tier 2 {tw[1]:.0f}%  Tier 3 {tw[2]:.0f}%")
    print()
else:
    CITY_WEIGHTS = [5] * 8 + [3] * 14 + [1] * 13   # Tier 1 gets more volume

PRODUCTS      = ["Personal", "Home", "Auto", "Business"]
EMP_TYPES     = ["Salaried", "Self-Employed", "Business Owner"]
LEAD_SOURCES  = ["Online", "Branch", "DSA", "Referral", "Walk-in"]
EMPLOYER_CATS = ["Government", "PSU", "Private", "MNC"]

TENURE_MAP = {
    "Personal": [12, 24, 36, 48, 60],
    "Home":     [60, 84, 120, 180, 240],
    "Auto":     [12, 24, 36, 48, 60],
    "Business": [12, 24, 36, 48, 60],
}

# ── GENERATE RECORDS ──────────────────────────────────────────────────────────
today      = datetime(2026, 3, 16)
start_date = today - timedelta(days=365)

rows = []
for i in range(1, NUM + 1):
    app_id = f"APP-{i:04d}"

    age    = random.randint(21, 65)
    gender = random.choices(["Male", "Female"], weights=[65, 35])[0]
    city, city_tier, state = random.choices(CITY_DATA, weights=CITY_WEIGHTS)[0]

    emp_type = random.choices(EMP_TYPES, weights=EMP_WEIGHTS)[0]
    if emp_type == "Salaried":
        employer_cat   = random.choice(EMPLOYER_CATS)
        monthly_income = int(np.random.randint(25_000, 200_001))
    elif emp_type == "Self-Employed":
        employer_cat   = "NA"
        monthly_income = int(np.random.randint(20_000, 300_001))
    else:  # Business Owner
        employer_cat   = "NA"
        monthly_income = int(np.random.randint(50_000, 500_001))

    existing_emi = round(monthly_income * np.random.uniform(0, 0.40), 2)
    cibil        = int(np.random.randint(600, 901))

    product = random.choices(PRODUCTS, weights=PRODUCT_WEIGHTS)[0]
    if product == "Home":
        loan_amt = int(np.random.randint(AMT_HOME[0],     AMT_HOME[1]     + 1))
    elif product == "Business":
        loan_amt = int(np.random.randint(AMT_BUSINESS[0], AMT_BUSINESS[1] + 1))
    elif product == "Auto":
        loan_amt = int(np.random.randint(AMT_AUTO[0],     AMT_AUTO[1]     + 1))
    else:
        loan_amt = int(np.random.randint(AMT_PERSONAL[0], AMT_PERSONAL[1] + 1))

    tenure      = random.choice(TENURE_MAP[product])
    lead_source = random.choices(LEAD_SOURCES, weights=[35, 20, 25, 15, 5])[0]
    app_date    = start_date + timedelta(days=random.randint(0, 365))
    foir        = round(existing_emi / monthly_income, 4) if monthly_income > 0 else 0

    rows.append([
        app_id, age, gender, city, city_tier, state,
        emp_type, employer_cat, monthly_income, round(existing_emi, 2),
        cibil, product, loan_amt, tenure,
        lead_source, app_date.strftime("%Y-%m-%d"), foir,
    ])

COLS = [
    "App_ID", "Age", "Gender", "City", "City_Tier", "State",
    "Employment_Type", "Employer_Category", "Monthly_Income", "Existing_EMI",
    "CIBIL_Score", "Loan_Product", "Loan_Amount_Requested", "Loan_Tenure_Months",
    "Lead_Source", "Application_Date", "FOIR",
]
df = pd.DataFrame(rows, columns=COLS)

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY       = PatternFill("solid", fgColor="1F4E79")
BLUE       = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
ALT        = PatternFill("solid", fgColor="EBF5FB")
WHITE      = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
KPI_FONT   = Font(bold=True, color="1F4E79", size=12)
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)

THIN_SIDE   = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def hdr(ws, row, col, value, fill=NAVY, font=HDR_FONT, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = fill, font, align, THIN_BORDER
    return c


def val(ws, row, col, value, fmt=None, fill=WHITE, align=RIGHT):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.fill, c.alignment, c.border = fill, align, THIN_BORDER
    return c


def section_title(ws, row, title, col_start=1, col_end=6):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.fill      = NAVY
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = THIN_BORDER
    ws.row_dimensions[row].height = 22


# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ── APPLICATIONS SHEET ────────────────────────────────────────────────────────
ws_a = wb.active
ws_a.title = "Applications"

for ci, name in enumerate(COLS, 1):
    hdr(ws_a, 1, ci, name)

col_fmts = {9: "#,##0.00", 10: "#,##0.00", 13: "#,##0.00", 17: "0.00%"}
for ri, row_data in enumerate(df.itertuples(index=False), 2):
    for ci, v in enumerate(row_data, 1):
        c = ws_a.cell(row=ri, column=ci, value=v)
        c.border = THIN_BORDER
        if ci in col_fmts:
            c.number_format = col_fmts[ci]

col_widths_a = [10, 5, 8, 16, 8, 16, 16, 16, 16, 14, 11, 10, 22, 16, 12, 18, 7]
for i, w in enumerate(col_widths_a, 1):
    ws_a.column_dimensions[get_column_letter(i)].width = w
ws_a.freeze_panes = "A2"

# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

for col, w in {1: 26, 2: 16, 3: 16, 4: 16, 5: 18, 6: 16}.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# Formula range aliases
last_row = NUM + 1

def rng(col):
    return f"Applications!${col}$2:${col}${last_row}"

EMP    = rng("G")
INC    = rng("I")
EMI    = rng("J")
CIB    = rng("K")
PROD   = rng("L")
LOAN   = rng("M")
TEN    = rng("N")
SRC    = rng("O")
FOIR_R = rng("Q")
TIER   = rng("E")
APP_ID = rng("A")

TOTAL  = f"=COUNTA({APP_ID})"   # reusable total count

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value     = "LOAN APPLICATION SUMMARY"
c.font      = TITLE_FONT
c.fill      = LIGHT_BLUE
c.alignment = CENTER
c.border    = THIN_BORDER
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
c = ws["A2"]
c.value     = f"Generated: {today.strftime('%d %b %Y')} | Total Records: {NUM:,}"
c.font      = Font(italic=True, color="666666", size=9)
c.fill      = WHITE
c.alignment = CENTER
c.border    = THIN_BORDER

R = 4

# ── SECTION 1: Executive KPIs ─────────────────────────────────────────────────
section_title(ws, R, "  EXECUTIVE KPIs")
R += 1

kpis = [
    ("Total Applications",      f"=COUNTA({APP_ID})",                   "0"),
    ("Total Loan Ask (Rs)",     f"=SUM({LOAN})",                        "#,##0.00"),
    ("Avg Loan Requested (Rs)", f"=AVERAGE({LOAN})",                    "#,##0.00"),
    ("Avg CIBIL Score",         f"=AVERAGE({CIB})",                     "0.0"),
    ("Avg Monthly Income (Rs)", f"=AVERAGE({INC})",                     "#,##0.00"),
    ("Avg Existing EMI (Rs)",   f"=AVERAGE({EMI})",                     "#,##0.00"),
    ("Avg FOIR",                f"=AVERAGE({FOIR_R})",                  "0.00%"),
    ("Avg Loan Tenure (Months)",f"=AVERAGE({TEN})",                     "0.0"),
    ("Salaried Applicants",     f'=COUNTIF({EMP},"Salaried")',          "0"),
    ("Online Applications",     f'=COUNTIF({SRC},"Online")',            "0"),
]

for idx, (label, formula, fmt) in enumerate(kpis):
    row = R + (idx // 2)
    col = 1 if idx % 2 == 0 else 4

    lc = ws.cell(row=row, column=col, value=label)
    lc.font, lc.fill, lc.alignment, lc.border = LABEL_FONT, LIGHT_BLUE, LEFT, THIN_BORDER

    ws.merge_cells(start_row=row, start_column=col+1,
                   end_row=row,   end_column=col+2)
    vc = ws.cell(row=row, column=col+1, value=formula)
    vc.font, vc.fill, vc.number_format = KPI_FONT, WHITE, fmt
    vc.alignment, vc.border = CENTER, THIN_BORDER

R += (len(kpis) + 1) // 2 + 1

# ── SECTION 2: Applications by Loan Product ───────────────────────────────────
section_title(ws, R, "  APPLICATIONS BY LOAN PRODUCT")
R += 1

for ci, h in enumerate(["Product", "Count", "% of Total",
                         "Avg CIBIL", "Avg Loan Ask (Rs)", "Avg FOIR"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, prod in enumerate(PRODUCTS):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, prod,                                                         fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({PROD},"{prod}")',              "#,##0",           fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({PROD},"{prod}")'
                  f'/COUNTA({APP_ID}),0)',                   "0.00%",           fill=fill)
    val(ws, R, 4, f'=AVERAGEIF({PROD},"{prod}",{CIB})',      "0.0",             fill=fill)
    val(ws, R, 5, f'=AVERAGEIF({PROD},"{prod}",{LOAN})',     "#,##0.00",        fill=fill)
    val(ws, R, 6, f'=AVERAGEIF({PROD},"{prod}",{FOIR_R})',   "0.00%",           fill=fill)
    R += 1

R += 1

# ── SECTION 3: Applications by Employment Type ────────────────────────────────
section_title(ws, R, "  APPLICATIONS BY EMPLOYMENT TYPE")
R += 1

for ci, h in enumerate(["Employment Type", "Count", "% of Total",
                         "Avg Income (Rs)", "Avg CIBIL", "Avg Loan Ask (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, emp in enumerate(EMP_TYPES):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, emp,                                                          fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({EMP},"{emp}")',                "#,##0",          fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({EMP},"{emp}")'
                  f'/COUNTA({APP_ID}),0)',                   "0.00%",          fill=fill)
    val(ws, R, 4, f'=AVERAGEIF({EMP},"{emp}",{INC})',        "#,##0.00",       fill=fill)
    val(ws, R, 5, f'=AVERAGEIF({EMP},"{emp}",{CIB})',        "0.0",            fill=fill)
    val(ws, R, 6, f'=AVERAGEIF({EMP},"{emp}",{LOAN})',       "#,##0.00",       fill=fill)
    R += 1

R += 1

# ── SECTION 4: Applications by City Tier ─────────────────────────────────────
section_title(ws, R, "  APPLICATIONS BY CITY TIER")
R += 1

for ci, h in enumerate(["City Tier", "Count", "% of Total",
                         "Avg Loan Ask (Rs)", "Avg CIBIL", "Avg Income (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, tier in enumerate(["Tier 1", "Tier 2", "Tier 3"]):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, tier,                                                         fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({TIER},"{tier}")',              "#,##0",          fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({TIER},"{tier}")'
                  f'/COUNTA({APP_ID}),0)',                   "0.00%",          fill=fill)
    val(ws, R, 4, f'=AVERAGEIF({TIER},"{tier}",{LOAN})',     "#,##0.00",       fill=fill)
    val(ws, R, 5, f'=AVERAGEIF({TIER},"{tier}",{CIB})',      "0.0",            fill=fill)
    val(ws, R, 6, f'=AVERAGEIF({TIER},"{tier}",{INC})',      "#,##0.00",       fill=fill)
    R += 1

R += 1

# ── SECTION 5: CIBIL Band Analysis ───────────────────────────────────────────
section_title(ws, R, "  CIBIL BAND ANALYSIS")
R += 1

for ci, h in enumerate(["CIBIL Band", "Count", "% of Total",
                         "Avg Loan Ask (Rs)", "Avg FOIR", "Avg Income (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

cibil_bands = [
    ("600 - 649", 600, 649),
    ("650 - 699", 650, 699),
    ("700 - 749", 700, 749),
    ("750 - 799", 750, 799),
    ("800 - 900", 800, 900),
]

for i, (label, lo, hi) in enumerate(cibil_bands):
    fill = ALT if i % 2 == 0 else WHITE
    mask = f"({CIB}>={lo})*({CIB}<={hi})"
    val(ws, R, 1, label,                                                                      fill=fill, align=LEFT)
    val(ws, R, 2, f"=SUMPRODUCT({mask}*1)",                                  "#,##0",        fill=fill)
    val(ws, R, 3, f"=IFERROR(SUMPRODUCT({mask}*1)/COUNTA({APP_ID}),0)",      "0.00%",        fill=fill)
    val(ws, R, 4, f"=IFERROR(SUMPRODUCT({mask}*{LOAN})/SUMPRODUCT({mask}*1),0)", "#,##0.00", fill=fill)
    val(ws, R, 5, f"=IFERROR(SUMPRODUCT({mask}*{FOIR_R})/SUMPRODUCT({mask}*1),0)", "0.00%", fill=fill)
    val(ws, R, 6, f"=IFERROR(SUMPRODUCT({mask}*{INC})/SUMPRODUCT({mask}*1),0)",  "#,##0.00", fill=fill)
    R += 1

R += 1

# ── SECTION 6: Lead Source Analysis ──────────────────────────────────────────
section_title(ws, R, "  LEAD SOURCE ANALYSIS")
R += 1

for ci, h in enumerate(["Lead Source", "Count", "% of Total",
                         "Avg CIBIL", "Avg Loan Ask (Rs)", "Avg FOIR"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, src in enumerate(LEAD_SOURCES):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, src,                                                          fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({SRC},"{src}")',                "#,##0",          fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({SRC},"{src}")'
                  f'/COUNTA({APP_ID}),0)',                   "0.00%",          fill=fill)
    val(ws, R, 4, f'=AVERAGEIF({SRC},"{src}",{CIB})',        "0.0",            fill=fill)
    val(ws, R, 5, f'=AVERAGEIF({SRC},"{src}",{LOAN})',       "#,##0.00",       fill=fill)
    val(ws, R, 6, f'=AVERAGEIF({SRC},"{src}",{FOIR_R})',     "0.00%",          fill=fill)
    R += 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "loan_applications.xlsx")
wb.save(out_path)
print(f"Loan applications saved: {out_path}")
