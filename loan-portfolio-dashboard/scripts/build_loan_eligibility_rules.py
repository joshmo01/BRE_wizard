"""
Loan Eligibility Rules Generator
Generates a human-readable, editable Excel file of BRE loan eligibility rules.

User should review / update the Value column before running generate_approved_loans.py.
The approved loans generator reads all rule parameters from this file.

Usage:
    python build_loan_eligibility_rules.py
    python build_loan_eligibility_rules.py --output "C:/path/to/folder"
"""

import argparse
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--output", default=r"C:\Users\joshm\OneDrive\Documents\BRE",
                    help="Output directory")
args = parser.parse_args()
OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)

# ── DEFAULT RULES ─────────────────────────────────────────────────────────────
# Columns: Rule_ID | Rule_Name | Field | Operator | Value | Applies_To | Enabled | Description
# Operator options : =  >  <  >=  <=  IN  NOT_IN
# Value for IN / NOT_IN: comma-separated list  e.g.  "Personal,Home"
# Applies_To: All  |  Salaried  |  Non-Salaried  |  Self-Employed  |  Business Owner
# Enabled: Yes / No

RULES = [
    ("R001", "Loan Product",            "Loan_Product",           "IN",  "Personal",
     "All",                      "Yes",
     "Eligible loan products. Comma-separate for multiple (e.g. Personal,Home,Auto)"),

    ("R002", "CIBIL Score – Minimum",   "CIBIL_Score",            ">",   "750",
     "All",                      "Yes",
     "Applicant CIBIL score must be strictly above this value"),

    ("R003", "FOIR Cap – Salaried",     "FOIR",                   "<",   "0.20",
     "Salaried",                 "Yes",
     "Fixed Obligation-to-Income Ratio must be below this for Salaried applicants (e.g. 0.20 = 20%)"),

    ("R004", "FOIR Cap – Non-Salaried", "FOIR",                   "<",   "0.15",
     "Self-Employed,Business Owner", "Yes",
     "FOIR cap for Self-Employed and Business Owner applicants (e.g. 0.15 = 15%)"),

    ("R005", "Max Loan Amount (Rs)",    "Loan_Amount_Requested",  "<",   "1500000",
     "All",                      "Yes",
     "Maximum requested loan amount in Rs. Enter numeric value without commas"),

    ("R006", "City Tier",               "City_Tier",              "IN",  "Tier 1,Tier 2,Tier 3",
     "All",                      "Yes",
     "Eligible city tiers. Valid values: Tier 1, Tier 2, Tier 3"),

    ("R007", "Minimum Age",             "Age",                    ">=",  "25",
     "All",                      "Yes",
     "Minimum applicant age in years"),

    ("R008", "Maximum Age",             "Age",                    "<=",  "50",
     "All",                      "Yes",
     "Maximum applicant age in years"),
]

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY        = PatternFill("solid", fgColor="1F4E79")
BLUE        = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F0")
YELLOW      = PatternFill("solid", fgColor="FFF2CC")   # editable value cells
GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")   # enabled rows
GREY        = PatternFill("solid", fgColor="F2F2F2")   # disabled rows
WHITE       = PatternFill("solid", fgColor="FFFFFF")
RED_LIGHT   = PatternFill("solid", fgColor="FCE4D6")   # disabled marker

THIN_SIDE   = Side(style="thin",   color="AAAAAA")
MED_SIDE    = Side(style="medium", color="1F4E79")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MED_BORDER  = Border(left=MED_SIDE,  right=MED_SIDE,  top=MED_SIDE,  bottom=MED_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
NOTE_FONT  = Font(italic=True, color="595959", size=9)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
VAL_FONT   = Font(bold=True, color="7B3F00", size=10)   # editable value highlight


def _cell(ws, row, col, value, fill=WHITE, font=None, align=LEFT, fmt=None, border=THIN_BORDER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = fill
    c.alignment = align
    c.border    = border
    if font:
        c.font = font
    if fmt:
        c.number_format = fmt
    return c


# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — ELIGIBILITY RULES (editable)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Eligibility Rules"
ws.sheet_view.showGridLines = False

# Column widths
col_widths = {1: 8, 2: 28, 3: 26, 4: 12, 5: 36, 6: 32, 7: 10, 8: 52}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── Title ──────────────────────────────────────────────────────────────────
ws.merge_cells("A1:H1")
c = ws["A1"]
c.value     = "LOAN ELIGIBILITY RULES — BRE CONFIGURATION"
c.font      = TITLE_FONT
c.fill      = LIGHT_BLUE
c.alignment = CENTER
c.border    = THIN_BORDER
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = (f"Generated: {datetime.today().strftime('%d %b %Y')}  |  "
           "Edit the Value column (highlighted in amber) to match your credit policy.  "
           "Set Enabled = No to deactivate a rule.  Save & close before running generate_approved_loans.py.")
c.font      = NOTE_FONT
c.fill      = WHITE
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border    = THIN_BORDER
ws.row_dimensions[2].height = 28

# ── Column headers ─────────────────────────────────────────────────────────
HEADERS = ["Rule ID", "Rule Name", "Field", "Operator", "Value ✏",
           "Applies To", "Enabled", "Description"]
R = 4
for ci, h in enumerate(HEADERS, 1):
    c = _cell(ws, R, ci, h, fill=NAVY, font=HDR_FONT, align=CENTER)
ws.row_dimensions[R].height = 26
R += 1

# ── Rule rows ──────────────────────────────────────────────────────────────
for rule in RULES:
    rid, name, field, op, value, applies, enabled, desc = rule
    is_enabled = enabled.strip().lower() == "yes"
    row_fill   = GREEN_LIGHT if is_enabled else GREY
    en_fill    = GREEN_LIGHT if is_enabled else RED_LIGHT

    ws.row_dimensions[R].height = 28
    _cell(ws, R, 1, rid,     fill=row_fill, font=Font(bold=True, size=9), align=CENTER)
    _cell(ws, R, 2, name,    fill=row_fill, font=LABEL_FONT)
    _cell(ws, R, 3, field,   fill=row_fill, font=Font(size=9, color="444444"), align=CENTER)
    _cell(ws, R, 4, op,      fill=row_fill, font=Font(bold=True, size=10, color="2E75B6"), align=CENTER)
    _cell(ws, R, 5, value,   fill=YELLOW,   font=VAL_FONT)           # ← editable
    _cell(ws, R, 6, applies, fill=row_fill, font=Font(size=9))
    _cell(ws, R, 7, enabled, fill=en_fill,  font=Font(bold=True, size=10,
                                                       color="375623" if is_enabled else "9C0006"),
          align=CENTER)
    _cell(ws, R, 8, desc,    fill=WHITE,    font=NOTE_FONT)
    R += 1

# ── Legend ─────────────────────────────────────────────────────────────────
R += 1
ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=8)
c = ws.cell(row=R, column=1,
            value="HOW TO EDIT:  "
                  "• Value column (amber) — change thresholds/lists directly.  "
                  "• IN / NOT_IN rules — comma-separate values, no spaces around commas.  "
                  "• Enabled column — type 'Yes' to activate, 'No' to skip.  "
                  "• Do NOT change Rule ID, Field, or Operator columns.")
c.fill      = LIGHT_BLUE
c.font      = Font(italic=True, color="1F4E79", size=9)
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border    = THIN_BORDER
ws.row_dimensions[R].height = 32

ws.freeze_panes = "A5"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RULE SUMMARY (read-only reference)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Rule Summary")
ws2.sheet_view.showGridLines = False

for col, w in {1: 28, 2: 52, 3: 16, 4: 12}.items():
    ws2.column_dimensions[get_column_letter(col)].width = w

ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value     = "RULE SUMMARY — PLAIN ENGLISH"
c.font      = TITLE_FONT
c.fill      = LIGHT_BLUE
c.alignment = CENTER
c.border    = THIN_BORDER
ws2.row_dimensions[1].height = 34

PLAIN = [
    ("Loan Product",         "Only applications for the listed product type(s) are eligible",         "R001", "IN list"),
    ("CIBIL Score",          "Applicant must score strictly above the minimum CIBIL threshold",        "R002", "Numeric"),
    ("FOIR – Salaried",      "Monthly obligations (EMI) ÷ income must be below cap for salaried",     "R003", "Decimal"),
    ("FOIR – Non-Salaried",  "Monthly obligations ÷ income below a tighter cap for SEP/Business",     "R004", "Decimal"),
    ("Max Loan Amount",      "Requested loan must not exceed the ceiling (in Rs)",                    "R005", "Numeric"),
    ("City Tier",            "Only applications from the listed city tiers are processed",             "R006", "IN list"),
    ("Age – Minimum",        "Applicant must be at or above the minimum age",                         "R007", "Integer"),
    ("Age – Maximum",        "Applicant must not exceed the maximum age",                             "R008", "Integer"),
]

R2 = 3
for ci, h in enumerate(["Rule Name", "Plain-English Description", "Rule ID", "Value Type"], 1):
    c = ws2.cell(row=R2, column=ci, value=h)
    c.fill, c.font, c.alignment, c.border = NAVY, HDR_FONT, CENTER, THIN_BORDER
ws2.row_dimensions[R2].height = 26
R2 += 1

for i, (name, desc, rid, vtype) in enumerate(PLAIN):
    fill = LIGHT_BLUE if i % 2 == 0 else WHITE
    ws2.row_dimensions[R2].height = 24
    for ci, v in enumerate([name, desc, rid, vtype], 1):
        c = ws2.cell(row=R2, column=ci, value=v)
        c.fill, c.alignment, c.border = fill, LEFT, THIN_BORDER
        if ci == 1:
            c.font = LABEL_FONT
    R2 += 1

# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "Loan Eligibility Rules.xlsx")
wb.save(out_path)
print(f"Loan Eligibility Rules saved: {out_path}")
print(f"  Rules defined : {len(RULES)}")
print(f"  Sheets        : Eligibility Rules (editable), Rule Summary (reference)")
print()
print("  NEXT STEP: Open the file, review/adjust the Value column (amber),")
print("  then save and close before running generate_approved_loans.py.")
