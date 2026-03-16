"""
Loan Lifecycle Dashboard Generator
Reads all three data files generated today and builds a unified formula-driven dashboard.

Sources:
  loan_applications.xlsx   → Applications sheet  (17 cols, A-Q)
  approved_loans.xlsx      → Approved_Loans sheet (24 cols, A-X)
  loan_repayment_schedule.xlsx → Repayment_Schedule sheet (21 cols, A-U)

Output: loan_lifecycle_dashboard.xlsx
  Sheet 1: Dashboard   (6 formula-driven sections)
  Sheet 2: Applications
  Sheet 3: Approved_Loans
  Sheet 4: Repayment_Schedule

Usage:
  python generate_lifecycle_dashboard.py
  python generate_lifecycle_dashboard.py --output "C:/path/to/folder"
"""

import argparse
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--apps",   default=r"C:\Users\joshm\OneDrive\Documents\BRE\loan_applications.xlsx")
parser.add_argument("--loans",  default=r"C:\Users\joshm\OneDrive\Documents\BRE\approved_loans.xlsx")
parser.add_argument("--sched",  default=r"C:\Users\joshm\OneDrive\Documents\BRE\loan_repayment_schedule.xlsx")
parser.add_argument("--output", default=r"C:\Users\joshm\OneDrive\Documents\BRE")
args = parser.parse_args()

OUT_PATH = os.path.join(args.output, "loan_lifecycle_dashboard.xlsx")
os.makedirs(args.output, exist_ok=True)

# ── LOAD DATA ─────────────────────────────────────────────────────────────────
df_apps  = pd.read_excel(args.apps,  sheet_name="Applications")
df_loans = pd.read_excel(args.loans, sheet_name="Approved_Loans")
df_sched = pd.read_excel(args.sched, sheet_name="Repayment_Schedule")

LR_A = len(df_apps)  + 1   # last data row in Applications sheet
LR_L = len(df_loans) + 1   # last data row in Approved_Loans sheet
LR_S = len(df_sched) + 1   # last data row in Repayment_Schedule sheet

print(f"Applications:       {len(df_apps):,} rows  (up to row {LR_A})")
print(f"Approved Loans:     {len(df_loans):,} rows  (up to row {LR_L})")
print(f"Repayment Schedule: {len(df_sched):,} rows  (up to row {LR_S})")

# ── WRITE DATA SHEETS ─────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

def write_df(wb, df, sheet_name):
    ws = wb.create_sheet(sheet_name)
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.font, cell.fill = hdr_font, hdr_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(c)].width = max(12, len(str(col)) + 2)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    ws.freeze_panes = "A2"
    return ws

write_df(wb, df_apps,  "Applications")
write_df(wb, df_loans, "Approved_Loans")
write_df(wb, df_sched, "Repayment_Schedule")

# ── CREATE DASHBOARD SHEET (insert at front) ──────────────────────────────────
ws = wb.create_sheet("Dashboard", 0)

# ── RANGE HELPERS ─────────────────────────────────────────────────────────────
def rng_a(col):  return f"Applications!${col}$2:${col}${LR_A}"
def rng_l(col):  return f"Approved_Loans!${col}$2:${col}${LR_L}"
def rng_s(col):  return f"Repayment_Schedule!${col}$2:${col}${LR_S}"

# ── STYLE CONSTANTS ───────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
TEAL       = "1F6B75"
GREEN      = "375623"
LIGHT_BLUE = "BDD7EE"
ALT_ROW    = "EBF3FB"
WHITE      = "FFFFFF"

def hf(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def tb():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def section_header(row, col, text, span=7, bg=DARK_BLUE, size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Calibri", bold=True, size=size, color="FFFFFF")
    cell.fill      = hf(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = tb()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    ws.row_dimensions[row].height = 22

def col_header(row, col, text, bg=MID_BLUE):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    c.fill      = hf(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = tb()
    ws.row_dimensions[row].height = 30

def kpi_label(row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
    c.fill      = hf(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = tb()
    ws.row_dimensions[row].height = 20

def kpi_value(row, col, expr, fmt="#,##0"):
    c = ws.cell(row=row, column=col, value="=" + expr)
    c.font        = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    c.fill        = hf(WHITE)
    c.alignment   = Alignment(horizontal="right", vertical="center")
    c.border      = tb()
    c.number_format = fmt

def row_label(row, col, text, alt=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
    c.fill      = hf(ALT_ROW if alt else WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = tb()
    ws.row_dimensions[row].height = 18

def data_cell(row, col, expr, fmt="#,##0", alt=False):
    c = ws.cell(row=row, column=col, value="=" + expr)
    c.font        = Font(name="Calibri", size=9)
    c.fill        = hf(ALT_ROW if alt else WHITE)
    c.alignment   = Alignment(horizontal="right", vertical="center")
    c.border      = tb()
    c.number_format = fmt

# ── COLUMN WIDTHS ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 2     # left margin
ws.column_dimensions["B"].width = 28    # label
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 2     # right margin

ws.row_dimensions[1].height = 6

# ── TITLE ─────────────────────────────────────────────────────────────────────
section_header(2, 2, "LOAN LIFECYCLE DASHBOARD", span=7, bg=DARK_BLUE, size=14)
ws.row_dimensions[2].height = 32

# ── SUBTITLE ROW (source labels) ──────────────────────────────────────────────
for c, (text, bg) in enumerate([
    ("Applications: loan_applications.xlsx", "2E75B6"),
    ("Approved Loans: approved_loans.xlsx",  "1F6B75"),
    ("Repayment: loan_repayment_schedule.xlsx", "375623"),
], 2):
    cell = ws.cell(row=3, column=c, value=text)
    cell.font      = Font(name="Calibri", size=8, color="FFFFFF", italic=True)
    cell.fill      = hf(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = tb()
ws.row_dimensions[3].height = 16

# =============================================================================
# SECTION 1 — LIFECYCLE KPIs  (rows 5-11)
# =============================================================================
section_header(5, 2, "1. LIFECYCLE KPIs", span=7, bg=MID_BLUE)

total_apps_expr  = f"COUNTA({rng_a('A')})"
total_loans_expr = f"COUNTA({rng_l('A')})"

kpis_left = [
    ("Total Applications",      total_apps_expr,                                           "#,##0"),
    ("Total Approved Loans",    total_loans_expr,                                          "#,##0"),
    ("Approval Rate %",         f"{total_loans_expr}/{total_apps_expr}*100",               "0.00\"%\""),
    ("Total Sanctioned (Rs)",   f"SUM({rng_l('O')})",                                      "#,##0"),
    ("Avg Interest Rate",       f"AVERAGE({rng_l('V')})*100",                              "0.00\"%\""),
]
kpis_right = [
    ("Total EMI Rows",           f"COUNTA({rng_s('A')})",                                  "#,##0"),
    ("Collection Efficiency %",  f"SUM({rng_s('U')})/SUM({rng_s('T')})*100",               "0.00\"%\""),
    ("Total Penal Interest (Rs)",f"SUM({rng_s('L')})",                                     "#,##0"),
    ("Total Default Penalty (Rs)",f"SUM({rng_s('P')})",                                    "#,##0"),
    ("Avg CIBIL (Approved)",     f"AVERAGE({rng_l('L')})",                                 "0.0"),
]

for i, (label, expr, fmt) in enumerate(kpis_left):
    r = 6 + i
    kpi_label(r, 2, label)
    kpi_value(r, 3, expr, fmt)

for i, (label, expr, fmt) in enumerate(kpis_right):
    r = 6 + i
    kpi_label(r, 5, label)
    kpi_value(r, 6, expr, fmt)

# =============================================================================
# SECTION 2 — APPLICATIONS BY LEAD SOURCE  (rows 13-21)
# Applications: O=Lead_Source, A=App_ID, K=CIBIL_Score, M=Loan_Amount_Requested, Q=FOIR
# =============================================================================
SEC2 = 13
section_header(SEC2, 2, "2. APPLICATIONS BY LEAD SOURCE", span=7, bg=MID_BLUE)
for c, h in enumerate(["Lead Source", "App Count", "% of Total",
                        "Avg CIBIL", "Avg Loan Ask (Rs)", "Avg FOIR %"], 2):
    col_header(SEC2 + 1, c, h)

total_apps = f"COUNTA({rng_a('A')})"
for i, src in enumerate(["Online", "Branch", "DSA", "Referral", "Walk-in"]):
    r, alt = SEC2 + 2 + i, i % 2 == 1
    row_label(r, 2, src, alt)
    cnt = f"COUNTIF({rng_a('O')},\"{src}\")"
    data_cell(r, 3, cnt,                                                        "#,##0",     alt)
    data_cell(r, 4, f"{cnt}/{total_apps}*100",                                  "0.00\"%\"", alt)
    data_cell(r, 5, f"AVERAGEIF({rng_a('O')},\"{src}\",{rng_a('K')})",          "0.0",       alt)
    data_cell(r, 6, f"AVERAGEIF({rng_a('O')},\"{src}\",{rng_a('M')})",          "#,##0",     alt)
    data_cell(r, 7, f"AVERAGEIF({rng_a('O')},\"{src}\",{rng_a('Q')})*100",      "0.00\"%\"", alt)

# =============================================================================
# SECTION 3 — APPROVED LOANS BY EMPLOYMENT TYPE  (rows 21-27)
# Approved_Loans: H=Employment_Type, L=CIBIL_Score, O=Sanctioned_Amount,
#                 V=Interest_Rate, W=EMI
# =============================================================================
SEC3 = SEC2 + 5 + 3
section_header(SEC3, 2, "3. APPROVED LOANS BY EMPLOYMENT TYPE", span=7, bg=TEAL)
for c, h in enumerate(["Employment Type", "Count", "% of Approved",
                        "Avg CIBIL", "Total Sanctioned (Rs)", "Avg Interest Rate"], 2):
    col_header(SEC3 + 1, c, h, bg=TEAL)

total_loans = f"COUNTA({rng_l('A')})"
for i, emp in enumerate(["Salaried", "Self-Employed", "Business Owner"]):
    r, alt = SEC3 + 2 + i, i % 2 == 1
    row_label(r, 2, emp, alt)
    cnt = f"COUNTIF({rng_l('H')},\"{emp}\")"
    data_cell(r, 3, cnt,                                                            "#,##0",     alt)
    data_cell(r, 4, f"{cnt}/{total_loans}*100",                                     "0.00\"%\"", alt)
    data_cell(r, 5, f"AVERAGEIF({rng_l('H')},\"{emp}\",{rng_l('L')})",              "0.0",       alt)
    data_cell(r, 6, f"SUMIF({rng_l('H')},\"{emp}\",{rng_l('O')})",                  "#,##0",     alt)
    data_cell(r, 7, f"AVERAGEIF({rng_l('H')},\"{emp}\",{rng_l('V')})*100",          "0.00\"%\"", alt)

# =============================================================================
# SECTION 4 — APPROVED LOANS BY CITY TIER  (rows ~30-35)
# Approved_Loans: F=City_Tier, L=CIBIL_Score, O=Sanctioned_Amount, W=EMI
# =============================================================================
SEC4 = SEC3 + 3 + 3
section_header(SEC4, 2, "4. APPROVED LOANS BY CITY TIER", span=7, bg=TEAL)
for c, h in enumerate(["City Tier", "Count", "% of Approved",
                        "Avg CIBIL", "Total Sanctioned (Rs)", "Avg EMI (Rs)"], 2):
    col_header(SEC4 + 1, c, h, bg=TEAL)

for i, tier in enumerate(["Tier 1", "Tier 2"]):
    r, alt = SEC4 + 2 + i, i % 2 == 1
    row_label(r, 2, tier, alt)
    cnt = f"COUNTIF({rng_l('F')},\"{tier}\")"
    data_cell(r, 3, cnt,                                                            "#,##0",     alt)
    data_cell(r, 4, f"{cnt}/{total_loans}*100",                                     "0.00\"%\"", alt)
    data_cell(r, 5, f"AVERAGEIF({rng_l('F')},\"{tier}\",{rng_l('L')})",             "0.0",       alt)
    data_cell(r, 6, f"SUMIF({rng_l('F')},\"{tier}\",{rng_l('O')})",                 "#,##0",     alt)
    data_cell(r, 7, f"AVERAGEIF({rng_l('F')},\"{tier}\",{rng_l('W')})",             "#,##0",     alt)

# =============================================================================
# SECTION 5 — CIBIL BAND ANALYSIS (APPROVED)  (rows ~36-43)
# Approved_Loans: L=CIBIL_Score, O=Sanctioned_Amount, V=Interest_Rate, W=EMI
# CIBIL bands for approved loans: 751-775, 776-800, 801-825, 826-850, 851-900
# =============================================================================
SEC5 = SEC4 + 2 + 3
section_header(SEC5, 2, "5. CIBIL BAND ANALYSIS (APPROVED LOANS)", span=7, bg=TEAL)
for c, h in enumerate(["CIBIL Band", "Count", "% of Approved",
                        "Avg Interest Rate", "Total Sanctioned (Rs)", "Avg EMI (Rs)"], 2):
    col_header(SEC5 + 1, c, h, bg=TEAL)

cibil_bands = [
    ("751 - 775",  751,  775),
    ("776 - 800",  776,  800),
    ("801 - 825",  801,  825),
    ("826 - 850",  826,  850),
    ("851 - 900",  851,  900),
]
for i, (label, lo, hi) in enumerate(cibil_bands):
    r, alt = SEC5 + 2 + i, i % 2 == 1
    row_label(r, 2, label, alt)
    cibil_r = rng_l('L')
    cnt = f"COUNTIFS({cibil_r},\">={lo}\",{cibil_r},\"<={hi}\")"
    data_cell(r, 3, cnt,                                                            "#,##0",     alt)
    data_cell(r, 4, f"{cnt}/{total_loans}*100",                                     "0.00\"%\"", alt)
    data_cell(r, 5,
              f"AVERAGEIFS({rng_l('V')},{cibil_r},\">={lo}\",{cibil_r},\"<={hi}\")*100",
              "0.00\"%\"", alt)
    data_cell(r, 6,
              f"SUMIFS({rng_l('O')},{cibil_r},\">={lo}\",{cibil_r},\"<={hi}\")",
              "#,##0",     alt)
    data_cell(r, 7,
              f"AVERAGEIFS({rng_l('W')},{cibil_r},\">={lo}\",{cibil_r},\"<={hi}\")",
              "#,##0",     alt)

# =============================================================================
# SECTION 6 — REPAYMENT PERFORMANCE  (rows ~45-52)
# Repayment_Schedule: S=Payment_Status, F=DPD, U=Total_Amount_Paid,
#                     L=Penal_Interest, G=DPD_Bucket
# =============================================================================
SEC6 = SEC5 + 5 + 3
section_header(SEC6, 2, "6. REPAYMENT PERFORMANCE BY PAYMENT STATUS", span=7, bg=GREEN)
for c, h in enumerate(["Payment Status", "EMI Count", "% of Total",
                        "Total Paid (Rs)", "Avg DPD", "Penal Interest (Rs)"], 2):
    col_header(SEC6 + 1, c, h, bg=GREEN)

total_emis = f"COUNTA({rng_s('A')})"
for i, status in enumerate(["On Time", "Delayed", "Defaulted", "Part Prepay", "Full Prepay"]):
    r, alt = SEC6 + 2 + i, i % 2 == 1
    row_label(r, 2, status, alt)
    cnt = f"COUNTIF({rng_s('S')},\"{status}\")"
    data_cell(r, 3, cnt,                                                            "#,##0",     alt)
    data_cell(r, 4, f"{cnt}/{total_emis}*100",                                      "0.00\"%\"", alt)
    data_cell(r, 5, f"SUMIF({rng_s('S')},\"{status}\",{rng_s('U')})",               "#,##0",     alt)
    data_cell(r, 6, f"AVERAGEIF({rng_s('S')},\"{status}\",{rng_s('F')})",           "0.0",       alt)
    data_cell(r, 7, f"SUMIF({rng_s('S')},\"{status}\",{rng_s('L')})",               "#,##0",     alt)

# ── FREEZE & ZOOM ─────────────────────────────────────────────────────────────
ws.freeze_panes = "B5"
ws.sheet_view.zoomScale = 90

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"Dashboard saved: {OUT_PATH}")
print(f"Sheets: {wb.sheetnames}")
