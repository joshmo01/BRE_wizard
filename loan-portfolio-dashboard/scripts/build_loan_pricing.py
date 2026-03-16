"""
Builds a complete loan pricing table with all combinations and suggested spreads.
Writes to: Loan Pricing.xlsx

Spread logic (additive over Base Rate of 9.00%):
  CIBIL score         → primary risk driver
  Employment type     → job stability premium
  Loan amount range   → larger loans slightly cheaper
  Tenure range        → longer tenure slightly more expensive
  Age group           → near-retirement small premium

Final Interest Rate = Base Rate + CIBIL Spread + Emp Spread + Amt Spread + Tenure Spread + Age Spread
"""

import itertools
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\joshm\OneDrive\Documents\BRE\Loan Pricing.xlsx"

# ── PRICING DIMENSIONS ────────────────────────────────────────────────────────
BASE_RATE = 9.00   # % — assumption; replace with MCLR/repo-linked rate

# (Employment Type, Employment Category)
EMP_COMBOS = [
    ("Salaried",      "Government"),
    ("Salaried",      "PSU"),
    ("Salaried",      "MNC"),
    ("Salaried",      "Private"),
    ("Self-Employed", "N/A"),
    ("Business Owner","N/A"),
]

PRODUCTS = ["Personal"]   # BRE currently approves only Personal Loans

CIBIL_BANDS = [
    ("751 - 775", 4.00),   # (label, CIBIL spread %)
    ("776 - 800", 2.75),
    ("801 - 825", 1.75),
    ("826 - 850", 1.00),
    ("851 - 900", 0.50),
]

AGE_GROUPS = [
    ("25 - 35", 0.00),
    ("36 - 45", 0.00),
    ("46 - 50", 0.25),
]

LOAN_AMOUNTS = [
    ("Up to 3L",  0.50),
    ("3L - 7L",   0.25),
    ("7L - 15L",  0.00),
]

TENURES = [
    ("12 - 24 months", -0.25),
    ("25 - 36 months",  0.00),
    ("37 - 60 months",  0.25),
]

# Employment spread
EMP_SPREAD = {
    ("Salaried",      "Government"): 0.00,
    ("Salaried",      "PSU"):        0.00,
    ("Salaried",      "MNC"):        0.25,
    ("Salaried",      "Private"):    0.50,
    ("Self-Employed", "N/A"):        0.75,
    ("Business Owner","N/A"):        1.00,
}

# ── BUILD ROWS ────────────────────────────────────────────────────────────────
rows = []
for (emp_type, emp_cat), product, (cibil_lbl, cibil_sp), \
    (age_lbl, age_sp), (amt_lbl, amt_sp), (tenure_lbl, tenure_sp) \
    in itertools.product(EMP_COMBOS, PRODUCTS, CIBIL_BANDS,
                         AGE_GROUPS, LOAN_AMOUNTS, TENURES):

    emp_sp    = EMP_SPREAD[(emp_type, emp_cat)]
    total_sp  = round(cibil_sp + emp_sp + amt_sp + tenure_sp + age_sp, 2)
    final_rate = round(BASE_RATE + total_sp, 2)

    rows.append({
        "Employment Type":     emp_type,
        "Employment Category": emp_cat,
        "Loan Product":        product,
        "CIBIL Score Range":   cibil_lbl,
        "Age Group":           age_lbl,
        "Loan Amount Range":   amt_lbl,
        "Tenure Range":        tenure_lbl,
        "CIBIL Spread (%)":    cibil_sp,
        "Emp Spread (%)":      emp_sp,
        "Amt Spread (%)":      amt_sp,
        "Tenure Spread (%)":   tenure_sp,
        "Age Spread (%)":      age_sp,
        "Total Spread (%)":    total_sp,
        "Final Rate (%)":      final_rate,
    })

df = pd.DataFrame(rows)
print(f"Total combinations: {len(df):,}")
print(f"Final rate range:   {df['Final Rate (%)'].min()}% — {df['Final Rate (%)'].max()}%")

# ── BUILD EXCEL ───────────────────────────────────────────────────────────────
wb  = Workbook()
ws  = wb.active
ws.title = "Pricing Table"

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
ORANGE     = "C55A11"
GREEN_DARK = "375623"
GREEN_MID  = "70AD47"
GREEN_LIGHT= "E2EFDA"
YELLOW     = "FFF2CC"
RED_LIGHT  = "FCE4D6"
WHITE      = "FFFFFF"
ALT        = "F2F8FF"

def hf(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def tb(color="B0B0B0"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("B2:O2")
t = ws["B2"]
t.value     = "PERSONAL LOAN PRICING TABLE — Spread Matrix"
t.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
t.fill      = hf(DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 28

ws.merge_cells("B3:O3")
sub = ws["B3"]
sub.value     = f"Base Rate: {BASE_RATE:.2f}%  |  Final Rate = Base Rate + Total Spread  |  All spreads in %"
sub.font      = Font(name="Calibri", size=9, italic=True, color=DARK_BLUE)
sub.fill      = hf(LIGHT_BLUE)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 16

# ── SPREAD LEGEND ─────────────────────────────────────────────────────────────
ws.merge_cells("B5:O5")
lg = ws["B5"]
lg.value     = "SPREAD COMPONENTS"
lg.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
lg.fill      = hf(MID_BLUE)
lg.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[5].height = 18

legend_items = [
    ("CIBIL 751-775", "4.00%"), ("CIBIL 776-800", "2.75%"), ("CIBIL 801-825", "1.75%"),
    ("CIBIL 826-850", "1.00%"), ("CIBIL 851-900", "0.50%"), None,
    ("Govt/PSU",      "0.00%"), ("MNC",           "0.25%"), ("Private",       "0.50%"),
    ("Self-Employed", "0.75%"), ("Business Owner", "1.00%"), None,
    ("Amt: Up to 3L", "+0.50%"), ("Amt: 3L-7L",   "+0.25%"), ("Amt: 7L-15L",  "0.00%"),
    ("Tenure: 12-24M","-0.25%"), ("Tenure: 25-36M","0.00%"), ("Tenure: 37-60M","+0.25%"),
    ("Age: 25-45",    "0.00%"), ("Age: 46-50",    "+0.25%"),
]

leg_row = 6
leg_col = 2
for item in legend_items:
    if item is None:
        leg_col += 1
        continue
    lbl, val = item
    lc = ws.cell(row=leg_row, column=leg_col, value=f"{lbl}: {val}")
    lc.font      = Font(name="Calibri", size=8)
    lc.fill      = hf(ALT)
    lc.alignment = Alignment(horizontal="left", indent=1)
    lc.border    = tb("D0D0D0")
    leg_col += 1
    if leg_col > 15:
        leg_col  = 2
        leg_row += 1
ws.row_dimensions[6].height = 14
ws.row_dimensions[7].height = 14

# ── HEADERS ───────────────────────────────────────────────────────────────────
COLS = list(df.columns)
HDR_ROW = 9

# Color-code header groups
HDR_COLORS = {
    "Employment Type":     DARK_BLUE,
    "Employment Category": DARK_BLUE,
    "Loan Product":        DARK_BLUE,
    "CIBIL Score Range":   MID_BLUE,
    "Age Group":           MID_BLUE,
    "Loan Amount Range":   MID_BLUE,
    "Tenure Range":        MID_BLUE,
    "CIBIL Spread (%)":    ORANGE,
    "Emp Spread (%)":      ORANGE,
    "Amt Spread (%)":      ORANGE,
    "Tenure Spread (%)":   ORANGE,
    "Age Spread (%)":      ORANGE,
    "Total Spread (%)":    GREEN_DARK,
    "Final Rate (%)":      GREEN_DARK,
}

for c, col in enumerate(COLS, 2):
    cell = ws.cell(row=HDR_ROW, column=c, value=col)
    cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    cell.fill      = hf(HDR_COLORS.get(col, DARK_BLUE))
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = tb()
ws.row_dimensions[HDR_ROW].height = 32

# ── DATA ROWS ─────────────────────────────────────────────────────────────────
for r_idx, row_data in enumerate(df.itertuples(index=False), HDR_ROW + 1):
    alt = r_idx % 2 == 0
    for c_idx, val in enumerate(row_data, 2):
        col_name = COLS[c_idx - 2]
        cell     = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.border = tb("D0D0D0")
        cell.font   = Font(name="Calibri", size=9)

        # Final Rate gets a colour-coded fill based on rate level
        if col_name == "Final Rate (%)":
            rate = float(val)
            if rate <= 11.0:
                cell.fill = hf(GREEN_LIGHT)
                cell.font = Font(name="Calibri", size=9, bold=True, color=GREEN_DARK)
            elif rate <= 13.0:
                cell.fill = hf(YELLOW)
                cell.font = Font(name="Calibri", size=9, bold=True, color="7F6000")
            else:
                cell.fill = hf(RED_LIGHT)
                cell.font = Font(name="Calibri", size=9, bold=True, color="843C0C")
            cell.alignment = Alignment(horizontal="center")
        elif col_name == "Total Spread (%)":
            cell.fill      = hf(ALT if alt else WHITE)
            cell.font      = Font(name="Calibri", size=9, bold=True, color=GREEN_DARK)
            cell.alignment = Alignment(horizontal="center")
        elif col_name in ("CIBIL Spread (%)", "Emp Spread (%)",
                          "Amt Spread (%)", "Tenure Spread (%)", "Age Spread (%)"):
            cell.fill      = hf(ALT if alt else WHITE)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.fill      = hf(ALT if alt else WHITE)
            cell.alignment = Alignment(horizontal="left", indent=1)

    ws.row_dimensions[r_idx].height = 15

# ── COLUMN WIDTHS ─────────────────────────────────────────────────────────────
COL_WIDTHS = {
    "B": 18, "C": 20, "D": 16, "E": 16, "F": 14,
    "G": 16, "H": 18, "I": 16, "J": 14, "K": 14,
    "L": 16, "M": 14, "N": 15, "O": 14,
}
for col_letter, width in COL_WIDTHS.items():
    ws.column_dimensions[col_letter].width = width

# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Rate Summary")

ws2.merge_cells("B2:H2")
sh = ws2["B2"]
sh.value     = "RATE SUMMARY — Avg Final Rate by CIBIL Band x Employment Category"
sh.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
sh.fill      = hf(DARK_BLUE)
sh.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[2].height = 24

pivot = df.pivot_table(
    index="CIBIL Score Range",
    columns="Employment Category",
    values="Final Rate (%)",
    aggfunc="mean"
).round(2)

cibil_order = ["751 - 775", "776 - 800", "801 - 825", "826 - 850", "851 - 900"]
emp_order   = ["Government", "PSU", "MNC", "Private", "N/A"]
pivot = pivot.reindex(index=cibil_order, columns=[c for c in emp_order if c in pivot.columns])

# Write pivot headers
ws2.cell(row=4, column=2, value="CIBIL Range").font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
ws2.cell(row=4, column=2).fill = hf(DARK_BLUE)
ws2.cell(row=4, column=2).border = tb()
ws2.row_dimensions[4].height = 22

for c_idx, col_name in enumerate(pivot.columns, 3):
    cell = ws2.cell(row=4, column=c_idx, value=col_name)
    cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    cell.fill      = hf(MID_BLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border    = tb()
    ws2.column_dimensions[get_column_letter(c_idx)].width = 16

ws2.column_dimensions["B"].width = 16

for r_idx, (cibil_lbl, row_vals) in enumerate(pivot.iterrows(), 5):
    ws2.row_dimensions[r_idx].height = 18
    lbl_cell = ws2.cell(row=r_idx, column=2, value=cibil_lbl)
    lbl_cell.font   = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
    lbl_cell.fill   = hf(LIGHT_BLUE)
    lbl_cell.border = tb()
    for c_idx, val in enumerate(row_vals, 3):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.border = tb()
        cell.alignment = Alignment(horizontal="center")
        if val <= 11.0:
            cell.fill = hf(GREEN_LIGHT)
            cell.font = Font(name="Calibri", size=9, bold=True, color=GREEN_DARK)
        elif val <= 13.0:
            cell.fill = hf(YELLOW)
            cell.font = Font(name="Calibri", size=9, bold=True, color="7F6000")
        else:
            cell.fill = hf(RED_LIGHT)
            cell.font = Font(name="Calibri", size=9, bold=True, color="843C0C")

ws2.freeze_panes = "C5"

# ── FREEZE & SAVE ─────────────────────────────────────────────────────────────
ws.freeze_panes = f"B{HDR_ROW + 1}"
ws.sheet_view.zoomScale = 85

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
print(f"\nRate distribution:")
print(df.groupby("Final Rate (%)")["Final Rate (%)"].count().rename("count").to_string())
