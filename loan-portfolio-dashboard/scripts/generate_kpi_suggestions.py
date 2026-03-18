"""
KPI Gap Analysis & Suggestions Generator
Compares Target KPIs vs Actual KPIs from the loan portfolio dashboard and generates
actionable suggestions across three levers:
  1. Funnel / Application Intake — expand or contract the customer pipeline
  2. Loan Pricing Matrix — adjust spreads, rates, or bands
  3. Rule Engine / Eligibility — tighten or relax BRE rules / collection policy

Reads:
  - loan_portfolio_dashboard.xlsx (Data sheet for actuals, Dashboard for KPIs)
  - Loan Pricing.xlsx (current pricing matrix)
  - Loan Eligibility Rules.xlsx (current BRE rules)
  - approved_loans.xlsx (for segment-level NPA/yield analysis)
  - loan_applications.xlsx (for funnel analysis)

Writes:
  - KPI_Gap_Suggestions.xlsx (3 sheets: Gap Summary, Detailed Suggestions, Data Analysis)

Usage:
    python generate_kpi_suggestions.py
    python generate_kpi_suggestions.py --output "C:/path/to/folder"
    python generate_kpi_suggestions.py --dashboard "C:/path/loan_portfolio_dashboard.xlsx" \
                                       --pricing "C:/path/Loan Pricing.xlsx" \
                                       --rules "C:/path/Loan Eligibility Rules.xlsx" \
                                       --approved "C:/path/approved_loans.xlsx" \
                                       --applications "C:/path/loan_applications.xlsx"
"""

import argparse
import os
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--output",       default=r"C:\Users\joshm\OneDrive\Documents\BRE",
                    help="Output directory")
parser.add_argument("--dashboard",    default="", help="Path to loan_portfolio_dashboard.xlsx")
parser.add_argument("--pricing",      default="", help="Path to Loan Pricing.xlsx")
parser.add_argument("--rules",        default="", help="Path to Loan Eligibility Rules.xlsx")
parser.add_argument("--approved",     default="", help="Path to approved_loans.xlsx")
parser.add_argument("--applications", default="", help="Path to loan_applications.xlsx")
args = parser.parse_args()

OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)


def _resolve(path, default_name):
    if path and os.path.isfile(path):
        return path
    candidate = os.path.join(OUT_DIR, default_name)
    return candidate if os.path.isfile(candidate) else None


DASH_PATH = _resolve(args.dashboard, "loan_portfolio_dashboard.xlsx")
PRICING_PATH = _resolve(args.pricing, "Loan Pricing.xlsx")
RULES_PATH = _resolve(args.rules, "Loan Eligibility Rules.xlsx")
APPROVED_PATH = _resolve(args.approved, "approved_loans.xlsx")
APPS_PATH = _resolve(args.applications, "loan_applications.xlsx")

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY        = PatternFill("solid", fgColor="1F4E79")
BLUE        = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F0")
ALT         = PatternFill("solid", fgColor="EBF5FB")
WHITE       = PatternFill("solid", fgColor="FFFFFF")
AMBER       = PatternFill("solid", fgColor="FFF2CC")
GREEN_RAG   = PatternFill("solid", fgColor="C6EFCE")
RED_RAG     = PatternFill("solid", fgColor="FFC7CE")
YELLOW_RAG  = PatternFill("solid", fgColor="FFEB9C")
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")
LIGHT_RED   = PatternFill("solid", fgColor="FCE4D6")

THIN_SIDE   = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")
WRAP   = Alignment(horizontal="left",   vertical="top", wrap_text=True)

TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
KPI_FONT   = Font(bold=True, color="1F4E79", size=12)
NOTE_FONT  = Font(italic=True, color="595959", size=9)
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(color="333333", size=10)
BOLD_BODY  = Font(bold=True, color="333333", size=10)
RED_FONT   = Font(bold=True, color="9C0006", size=10)
GREEN_FONT = Font(bold=True, color="375623", size=10)


def hdr(ws, row, col, value, fill=NAVY, font=HDR_FONT, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = fill, font, align, THIN_BORDER
    return c


def val(ws, row, col, value, fmt=None, fill=WHITE, align=LEFT, font=BODY_FONT):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.fill, c.alignment, c.border, c.font = fill, align, THIN_BORDER, font
    return c


def section_title(ws, row, title, col_end=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=1, value=title)
    c.fill, c.font = NAVY, Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


# ── LOAD DASHBOARD DATA ──────────────────────────────────────────────────────
print("[1/5] Loading dashboard data...")

if not DASH_PATH:
    print("ERROR: loan_portfolio_dashboard.xlsx not found. Run generate_dashboard.py first.")
    raise SystemExit(1)

df = pd.read_excel(DASH_PATH, sheet_name="Data")

# Compute actuals
act = {}
act["Total Loans"]             = len(df)
act["Total Disbursed"]         = df["Disbursement_Amt"].sum()
act["Total Outstanding"]       = df["Outstanding_Principal"].sum()
act["NPA Count"]               = int((df["Status"] == "NPA").sum())
act["NPA Rate"]                = act["NPA Count"] / len(df) if len(df) > 0 else 0
act["Delinquent Count"]        = int((df["Status"] == "Delinquent").sum())
act["Portfolio Yield"]         = ((df["Interest_Rate"] * df["Outstanding_Principal"]).sum()
                                  / df["Outstanding_Principal"].sum()) if df["Outstanding_Principal"].sum() > 0 else 0
act["Monthly Profit"]          = df["Monthly_Profit"].sum()
act["Avg Loan Size"]           = df["Outstanding_Principal"].sum() / len(df) if len(df) > 0 else 0
act["Active Loans"]            = int((df["Status"] == "Active").sum())

# Read target values from the Dashboard sheet (column C = Target, rows 6-15 typically)
dash_wb = load_workbook(DASH_PATH, data_only=True)
dash_ws = dash_wb["Dashboard"]

tgt = {}
kpi_names = [
    "Total Loans", "Total Disbursed", "Total Outstanding",
    "NPA Count", "NPA Rate", "Delinquent Count",
    "Portfolio Yield", "Monthly Profit", "Avg Loan Size", "Active Loans"
]

# Find KPI rows by scanning column A for known labels
for row in dash_ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=4):
    cell_a = row[0].value
    if cell_a and isinstance(cell_a, str):
        for kn in kpi_names:
            if kn.lower() in cell_a.lower() or kn.replace(" ", "").lower() in cell_a.replace(" ", "").lower():
                target_val = row[2].value  # column C = Target
                if target_val is not None and isinstance(target_val, (int, float)):
                    tgt[kn] = target_val
                break

# Fill missing targets with actuals (no gap)
for kn in kpi_names:
    if kn not in tgt:
        tgt[kn] = act[kn]

dash_wb.close()

# Direction: True = higher is better, False = lower is better
DIRECTION = {
    "Total Loans": True, "Total Disbursed": True, "Total Outstanding": True,
    "NPA Count": False, "NPA Rate": False, "Delinquent Count": False,
    "Portfolio Yield": True, "Monthly Profit": True, "Avg Loan Size": True,
    "Active Loans": True,
}

# ── SEGMENT-LEVEL ANALYSIS ───────────────────────────────────────────────────
print("[2/5] Analysing segment-level KPIs...")

segments = df["Segment"].unique().tolist()
products = df["Product"].unique().tolist()
ratings  = df["Credit_Rating"].unique().tolist()

# NPA by segment
seg_npa = df.groupby("Segment").apply(include_groups=False,func=
    lambda g: pd.Series({
        "Count": len(g),
        "NPA": int((g["Status"] == "NPA").sum()),
        "NPA%": (g["Status"] == "NPA").sum() / len(g) if len(g) > 0 else 0,
        "Outstanding": g["Outstanding_Principal"].sum(),
        "Yield": ((g["Interest_Rate"] * g["Outstanding_Principal"]).sum()
                  / g["Outstanding_Principal"].sum()) if g["Outstanding_Principal"].sum() > 0 else 0,
        "Avg_Rate": g["Interest_Rate"].mean(),
    })
).reset_index()

# NPA by credit rating
rat_npa = df.groupby("Credit_Rating").apply(include_groups=False,func=
    lambda g: pd.Series({
        "Count": len(g),
        "NPA": int((g["Status"] == "NPA").sum()),
        "NPA%": (g["Status"] == "NPA").sum() / len(g) if len(g) > 0 else 0,
        "Outstanding": g["Outstanding_Principal"].sum(),
        "Avg_Rate": g["Interest_Rate"].mean(),
        "Avg_Profit": g["Monthly_Profit"].mean(),
    })
).reset_index()

# NPA by product
prod_npa = df.groupby("Product").apply(include_groups=False,func=
    lambda g: pd.Series({
        "Count": len(g),
        "NPA": int((g["Status"] == "NPA").sum()),
        "NPA%": (g["Status"] == "NPA").sum() / len(g) if len(g) > 0 else 0,
        "Outstanding": g["Outstanding_Principal"].sum(),
        "Yield": ((g["Interest_Rate"] * g["Outstanding_Principal"]).sum()
                  / g["Outstanding_Principal"].sum()) if g["Outstanding_Principal"].sum() > 0 else 0,
    })
).reset_index()

# ── LOAD CURRENT RULES & PRICING ─────────────────────────────────────────────
print("[3/5] Loading current rules and pricing...")

current_rules = {}
if RULES_PATH:
    try:
        rules_df = pd.read_excel(RULES_PATH, sheet_name="Eligibility Rules")
        for _, r in rules_df.iterrows():
            rid = r.get("Rule_ID", r.get("Rule ID", ""))
            current_rules[rid] = {
                "name": r.get("Rule_Name", r.get("Rule Name", "")),
                "value": str(r.get("Value", "")),
                "enabled": str(r.get("Enabled", "Yes")),
                "operator": str(r.get("Operator", "")),
            }
    except Exception as e:
        print(f"  Warning: Could not read rules file: {e}")

current_pricing = {}
if PRICING_PATH:
    try:
        pricing_df = pd.read_excel(PRICING_PATH, sheet_name="Pricing Table")
        current_pricing["base_rate"] = 9.00  # default
        if "Final Rate (%)" in pricing_df.columns:
            current_pricing["min_rate"] = pricing_df["Final Rate (%)"].min()
            current_pricing["max_rate"] = pricing_df["Final Rate (%)"].max()
            current_pricing["avg_rate"] = pricing_df["Final Rate (%)"].mean()
        if "CIBIL Spread (%)" in pricing_df.columns:
            current_pricing["max_cibil_spread"] = pricing_df["CIBIL Spread (%)"].max()
            current_pricing["min_cibil_spread"] = pricing_df["CIBIL Spread (%)"].min()
    except Exception as e:
        print(f"  Warning: Could not read pricing file: {e}")

# Application funnel stats
app_stats = {}
if APPS_PATH:
    try:
        apps_df = pd.read_excel(APPS_PATH, sheet_name="Applications")
        app_stats["total_applications"] = len(apps_df)
        if "Employment_Type" in apps_df.columns:
            app_stats["emp_dist"] = apps_df["Employment_Type"].value_counts().to_dict()
        if "Loan_Product" in apps_df.columns:
            app_stats["product_dist"] = apps_df["Loan_Product"].value_counts().to_dict()
    except Exception as e:
        print(f"  Warning: Could not read applications file: {e}")

approval_stats = {}
if APPROVED_PATH:
    try:
        appr_df = pd.read_excel(APPROVED_PATH, sheet_name="Approved_Loans")
        approval_stats["total_approved"] = len(appr_df)
        if app_stats.get("total_applications"):
            approval_stats["approval_rate"] = len(appr_df) / app_stats["total_applications"]
    except Exception as e:
        print(f"  Warning: Could not read approved loans file: {e}")


# ── GENERATE SUGGESTIONS ─────────────────────────────────────────────────────
print("[4/5] Generating suggestions...")

# Each suggestion: (KPI, Gap Description, Lever, Suggestion, Priority, Current Value, Suggested Change)
suggestions = []


def gap_pct(actual, target, higher_is_better):
    """Return gap as a signed percentage. Negative = off target."""
    if target == 0:
        return 0
    gap = (actual - target) / abs(target) * 100
    return gap if higher_is_better else -gap


def add_suggestion(kpi, gap_desc, lever, suggestion, priority, current_val, suggested_change):
    suggestions.append({
        "KPI": kpi,
        "Gap": gap_desc,
        "Lever": lever,
        "Suggestion": suggestion,
        "Priority": priority,
        "Current Value": current_val,
        "Suggested Change": suggested_change,
    })


# ── 1. DISBURSEMENT / TOTAL LOANS GAP ────────────────────────────────────────
for kpi_name in ["Total Loans", "Total Disbursed", "Active Loans"]:
    g = gap_pct(act[kpi_name], tgt[kpi_name], True)
    if g < -5:  # off target by more than 5%
        shortfall = tgt[kpi_name] - act[kpi_name]
        pct_short = abs(g)

        # Funnel suggestions
        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Funnel / Application Intake",
            "Expand customer segments: Add SENP (Self-Employed Non-Professional) category to "
            "application intake. Currently only Salaried, Self-Employed, and Business Owner are "
            "considered. SENP (doctors, lawyers, CAs, architects) are typically lower-risk self-employed "
            "with stable income, and can meaningfully increase the funnel.",
            "High" if pct_short > 15 else "Medium",
            f"Current segments: Salaried, Self-Employed, Business Owner",
            f"Add: SENP (Self-Employed Non-Professional) — estimated +15-20% funnel increase",
        )

        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Funnel / Application Intake",
            "Expand eligible loan products. Currently only Personal Loans are approved by the BRE. "
            "Adding Home Loan or Auto Loan products can increase both the number of disbursements "
            "and the average loan size.",
            "High" if pct_short > 20 else "Medium",
            f"Current products: {current_rules.get('R001', {}).get('value', 'Personal')}",
            f"Add: Home, Auto (update R001 value to 'Personal,Home,Auto')",
        )

        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Rule Engine / Eligibility",
            "Relax CIBIL score threshold. Lowering the minimum CIBIL from the current cutoff "
            "to 700 would admit more applicants. This should be paired with risk-adjusted "
            "pricing (higher spreads for 700-750 band) to maintain portfolio quality.",
            "Medium",
            f"Current CIBIL min: {current_rules.get('R002', {}).get('value', '750')}",
            f"Suggested: Lower to 700 (with higher pricing spread for 700-750 band)",
        )

        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Rule Engine / Eligibility",
            "Increase FOIR cap for salaried employees. A higher FOIR cap (e.g. 25%) allows "
            "applicants with existing obligations to still qualify, expanding the eligible pool.",
            "Low",
            f"Current FOIR cap (salaried): {current_rules.get('R003', {}).get('value', '0.20')}",
            f"Suggested: Increase to 0.25 (25%)",
        )

        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Funnel / Application Intake",
            "Increase lead generation through additional channels. Add digital partnerships "
            "(fintech tie-ups, aggregator platforms), corporate salary account tie-ups, and "
            "targeted campaigns in Tier 2/3 cities where penetration is lower.",
            "Medium",
            f"Current lead sources: Online, Branch, DSA, Referral, Walk-in",
            f"Add: Fintech aggregators, Corporate tie-ups, WhatsApp/SMS campaigns",
        )

        # Age range
        add_suggestion(
            kpi_name,
            f"Actual {act[kpi_name]:,.0f} vs Target {tgt[kpi_name]:,.0f} ({pct_short:.1f}% below target)",
            "Rule Engine / Eligibility",
            "Widen the eligible age range. Reducing min age from 25 to 21 and increasing max "
            "age from 50 to 58 would capture younger professionals (IT, startup sector) and "
            "pre-retirement earners with strong repayment capacity.",
            "Low",
            f"Current: Min {current_rules.get('R007', {}).get('value', '25')}, "
            f"Max {current_rules.get('R008', {}).get('value', '50')}",
            f"Suggested: Min 21, Max 58",
        )

# ── 2. PORTFOLIO YIELD GAP ──────────────────────────────────────────────────
g_yield = gap_pct(act["Portfolio Yield"], tgt["Portfolio Yield"], True)
if g_yield < -2:  # yield below target
    pct_short = abs(g_yield)

    # Find low-yield segments
    low_yield_ratings = rat_npa[rat_npa["Avg_Rate"] < act["Portfolio Yield"]].sort_values("Avg_Rate")

    add_suggestion(
        "Portfolio Yield",
        f"Actual {act['Portfolio Yield']:.2%} vs Target {tgt['Portfolio Yield']:.2%} "
        f"({pct_short:.1f}% below target)",
        "Loan Pricing Matrix",
        "Increase spreads for lower CIBIL bands. The lower CIBIL bands (751-775, 776-800) "
        "represent higher risk but their pricing spread may not adequately compensate. "
        "Increase the CIBIL spread by 0.50-1.00% for these bands to improve portfolio yield.",
        "High",
        f"Current CIBIL spread range: {current_pricing.get('min_cibil_spread', 0.5):.2f}% - "
        f"{current_pricing.get('max_cibil_spread', 4.0):.2f}%",
        f"Increase bottom two CIBIL band spreads by +0.50% each",
    )

    add_suggestion(
        "Portfolio Yield",
        f"Actual {act['Portfolio Yield']:.2%} vs Target {tgt['Portfolio Yield']:.2%} "
        f"({pct_short:.1f}% below target)",
        "Loan Pricing Matrix",
        "Increase base rate or reduce the rate discount for high-ticket loans. "
        "Currently larger loans (7L-15L) have 0% amount spread — adding +0.25% "
        "for this band improves overall yield without affecting small-ticket borrowers.",
        "Medium",
        f"Current base rate: {current_pricing.get('base_rate', 9.0):.2f}%",
        f"Add +0.25% amount spread for 7L-15L band; or raise base rate by 0.25%",
    )

    add_suggestion(
        "Portfolio Yield",
        f"Actual {act['Portfolio Yield']:.2%} vs Target {tgt['Portfolio Yield']:.2%} "
        f"({pct_short:.1f}% below target)",
        "Loan Pricing Matrix",
        "Increase tenure spread for longer tenures. Loans with 37-60 month tenure "
        "carry higher duration risk. Increasing the tenure spread from +0.25% to +0.50% "
        "for this band improves yield on the longest-duration book.",
        "Low",
        f"Current long-tenure spread: +0.25%",
        f"Suggested: +0.50% for 37-60 month tenure",
    )

    # Suggest pricing adjustment for employment categories
    add_suggestion(
        "Portfolio Yield",
        f"Actual {act['Portfolio Yield']:.2%} vs Target {tgt['Portfolio Yield']:.2%} "
        f"({pct_short:.1f}% below target)",
        "Loan Pricing Matrix",
        "Review employment-category pricing. Private-sector salaried employees have higher "
        "attrition risk than Government/PSU. Consider increasing the employment spread for "
        "Private-sector from current level to +0.75% (from +0.50%) to reflect job-stability risk.",
        "Medium",
        f"Current Private spread: +0.50%",
        f"Suggested: +0.75% for Private sector salaried",
    )

# ── 3. NPA / DELINQUENCY GAP ─────────────────────────────────────────────────
for kpi_name, act_key in [("NPA Rate", "NPA Rate"), ("NPA Count", "NPA Count"),
                           ("Delinquent Count", "Delinquent Count")]:
    g = gap_pct(act[act_key], tgt[act_key], False)  # lower is better
    if g < -5:  # worse than target by >5%
        pct_over = abs(g)

        # Find worst NPA segments
        worst_seg = seg_npa.sort_values("NPA%", ascending=False)
        worst_seg_name = worst_seg.iloc[0]["Segment"] if len(worst_seg) > 0 else "Unknown"
        worst_seg_npa = worst_seg.iloc[0]["NPA%"] if len(worst_seg) > 0 else 0

        worst_rat = rat_npa.sort_values("NPA%", ascending=False)
        worst_rat_name = worst_rat.iloc[0]["Credit_Rating"] if len(worst_rat) > 0 else "Unknown"
        worst_rat_npa = worst_rat.iloc[0]["NPA%"] if len(worst_rat) > 0 else 0

        add_suggestion(
            kpi_name,
            f"Actual {act[act_key]:,.2f} vs Target {tgt[act_key]:,.2f} "
            f"({pct_over:.1f}% above target)" if isinstance(act[act_key], float) else
            f"Actual {act[act_key]:,} vs Target {tgt[act_key]:,.0f} "
            f"({pct_over:.1f}% above target)",
            "Rule Engine / Eligibility",
            f"Tighten eligibility for high-NPA segment: '{worst_seg_name}' has {worst_seg_npa:.1%} NPA rate. "
            f"Consider excluding or adding stricter criteria for this segment — e.g., higher CIBIL "
            f"threshold (+25 points), lower FOIR cap (-5%), or additional documentation requirements.",
            "High",
            f"Segment '{worst_seg_name}': NPA {worst_seg_npa:.1%}",
            f"Add segment-specific CIBIL floor: {worst_seg_name} requires CIBIL > 775",
        )

        add_suggestion(
            kpi_name,
            f"Actual {act[act_key]:,.2f} vs Target {tgt[act_key]:,.2f} "
            f"({pct_over:.1f}% above target)" if isinstance(act[act_key], float) else
            f"Actual {act[act_key]:,} vs Target {tgt[act_key]:,.0f} "
            f"({pct_over:.1f}% above target)",
            "Rule Engine / Eligibility",
            f"Tighten rules for credit rating '{worst_rat_name}' (NPA: {worst_rat_npa:.1%}). "
            f"Options: (a) Exclude this rating entirely, (b) Require collateral/guarantee, "
            f"(c) Cap loan amount at 50% of current max for this rating.",
            "High",
            f"Rating '{worst_rat_name}': NPA {worst_rat_npa:.1%}",
            f"Exclude rating '{worst_rat_name}' or cap loan amount to Rs 7,50,000",
        )

        add_suggestion(
            kpi_name,
            f"Actual is {pct_over:.1f}% above target",
            "Collection Policy",
            "Strengthen early-stage collection for 1-30 DPD bucket. Implement automated SMS/IVR "
            "reminders at DPD 1, 7, and 15. Assign telecaller follow-up at DPD 7 for accounts "
            "above Rs 5 lakhs outstanding. This prevents delinquent accounts from slipping into NPA.",
            "High",
            f"Current: Standard collection process",
            f"Add: Automated reminders (DPD 1, 7, 15) + telecaller at DPD 7 for > Rs 5L",
        )

        add_suggestion(
            kpi_name,
            f"Actual is {pct_over:.1f}% above target",
            "Collection Policy",
            "Implement risk-based collection intensity. High-risk accounts (lower CIBIL, higher "
            "outstanding) should get field visits at DPD 30 instead of DPD 60. Re-classify "
            "collection priority based on expected loss rather than just DPD bucket.",
            "Medium",
            f"Current: Uniform DPD-based collection",
            f"Move to: Risk-based collection — field visits at DPD 30 for high-risk",
        )

        add_suggestion(
            kpi_name,
            f"Actual is {pct_over:.1f}% above target",
            "Loan Pricing Matrix",
            "Increase risk premium for segments with high NPA. The pricing should reflect the "
            "actual credit cost. Add a segment-level spread of +0.50% for segments where "
            "NPA > 5% to build a loss buffer into the yield.",
            "Medium",
            f"Current: No segment-level spread",
            f"Add: +0.50% segment spread where NPA > 5%",
        )

# ── 4. PROFIT GAP ────────────────────────────────────────────────────────────
g_profit = gap_pct(act["Monthly Profit"], tgt["Monthly Profit"], True)
if g_profit < -5:
    pct_short = abs(g_profit)

    add_suggestion(
        "Monthly Profit",
        f"Actual Rs {act['Monthly Profit']:,.0f} vs Target Rs {tgt['Monthly Profit']:,.0f} "
        f"({pct_short:.1f}% below target)",
        "Loan Pricing Matrix",
        "Review cost-of-funds assumption. If the cost-of-funds (currently assumed 3% in dashboard) "
        "has increased, reprice the book accordingly. Every 0.25% increase in spread across the "
        f"portfolio adds approximately Rs {act['Total Outstanding'] * 0.0025 / 12:,.0f}/month to profit.",
        "High",
        f"Current estimated NIM: {(act['Portfolio Yield'] - 0.03) * 100:.2f}%",
        f"Target NIM needed: ~{((tgt['Monthly Profit'] * 12 / act['Total Outstanding']) + 0.03) * 100:.2f}% "
        f"(increase spread by ~{max(0, (tgt['Monthly Profit'] - act['Monthly Profit']) * 12 / act['Total Outstanding'] * 100):.2f}%)"
        if act["Total Outstanding"] > 0 else "Increase overall spread",
    )

    add_suggestion(
        "Monthly Profit",
        f"Actual Rs {act['Monthly Profit']:,.0f} vs Target Rs {tgt['Monthly Profit']:,.0f} "
        f"({pct_short:.1f}% below target)",
        "Funnel / Application Intake",
        "Increase portfolio size. Profit gap can be closed partially by growing the book. "
        "Combine with funnel expansion suggestions above to increase total outstanding "
        "while maintaining yield.",
        "Medium",
        f"Current outstanding: Rs {act['Total Outstanding']:,.0f}",
        f"Need ~Rs {(tgt['Monthly Profit'] * 12 / (act['Portfolio Yield'] - 0.03)):,.0f} outstanding "
        f"at current yield to hit profit target"
        if act["Portfolio Yield"] > 0.03 else "Grow outstanding",
    )

# ── 5. AVG LOAN SIZE GAP ─────────────────────────────────────────────────────
g_als = gap_pct(act["Avg Loan Size"], tgt["Avg Loan Size"], True)
if g_als < -5:
    add_suggestion(
        "Avg Loan Size",
        f"Actual Rs {act['Avg Loan Size']:,.0f} vs Target Rs {tgt['Avg Loan Size']:,.0f}",
        "Rule Engine / Eligibility",
        "Increase the maximum loan amount cap (R005). Raising from Rs 15L to Rs 25L allows "
        "high-creditworthy borrowers to take larger loans, pulling up the average loan size.",
        "Medium",
        f"Current max: Rs {current_rules.get('R005', {}).get('value', '1500000')}",
        f"Suggested: Rs 25,00,000",
    )

    add_suggestion(
        "Avg Loan Size",
        f"Actual Rs {act['Avg Loan Size']:,.0f} vs Target Rs {tgt['Avg Loan Size']:,.0f}",
        "Funnel / Application Intake",
        "Add Home Loan and Business Loan products which have naturally higher ticket sizes "
        "(Rs 10L-1Cr for Home, Rs 5L-50L for Business) compared to Personal Loans (Rs 50K-15L).",
        "High",
        f"Current products: {current_rules.get('R001', {}).get('value', 'Personal')}",
        f"Add: Home, Business (avg ticket size 5-10x higher)",
    )


# ── ADD ON-TRACK KPIs (no suggestions needed) ────────────────────────────────
# If no suggestions were generated, note that everything is on track
if not suggestions:
    add_suggestion(
        "All KPIs",
        "All KPIs are on track or within 5% of target",
        "None",
        "No immediate changes recommended. Continue monitoring monthly.",
        "Info",
        "All within tolerance",
        "No change required",
    )


# ── BUILD EXCEL OUTPUT ────────────────────────────────────────────────────────
print("[5/5] Writing KPI_Gap_Suggestions.xlsx...")

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: GAP SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Gap Summary"
ws1.sheet_view.showGridLines = False

# Title
ws1.merge_cells("A1:F1")
c = ws1["A1"]
c.value = "KPI GAP ANALYSIS — TARGET vs ACTUAL"
c.font, c.fill, c.alignment, c.border = TITLE_FONT, LIGHT_BLUE, CENTER, THIN_BORDER
ws1.row_dimensions[1].height = 36

ws1.merge_cells("A2:F2")
c = ws1["A2"]
c.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Source: loan_portfolio_dashboard.xlsx"
c.font, c.fill, c.alignment, c.border = NOTE_FONT, WHITE, CENTER, THIN_BORDER

R = 4
section_title(ws1, R, "  KPI COMPARISON — ACTUAL vs TARGET", col_end=6)
R += 1

headers = ["KPI", "Actual", "Target", "Gap %", "Status", "Direction"]
for ci, h in enumerate(headers, 1):
    hdr(ws1, R, ci, h, fill=BLUE, font=SUB_FONT)
ws1.row_dimensions[R].height = 24
R += 1

formats = {
    "Total Loans": "0", "Total Disbursed": "#,##0", "Total Outstanding": "#,##0",
    "NPA Count": "0", "NPA Rate": "0.00%", "Delinquent Count": "0",
    "Portfolio Yield": "0.00%", "Monthly Profit": "#,##0", "Avg Loan Size": "#,##0",
    "Active Loans": "0",
}

for kn in kpi_names:
    fill = ALT if (R % 2 == 0) else WHITE
    higher = DIRECTION[kn]
    g = gap_pct(act[kn], tgt[kn], higher)
    on_track = g >= -5

    val(ws1, R, 1, kn, fill=fill, font=LABEL_FONT)
    val(ws1, R, 2, act[kn], fmt=formats[kn], fill=fill, align=RIGHT)
    val(ws1, R, 3, tgt[kn], fmt=formats[kn], fill=AMBER, align=RIGHT, font=BOLD_BODY)
    val(ws1, R, 4, f"{g:+.1f}%", fill=GREEN_RAG if on_track else RED_RAG, align=CENTER,
        font=GREEN_FONT if on_track else RED_FONT)
    val(ws1, R, 5, "✓ On Track" if on_track else "✗ Off Target",
        fill=GREEN_RAG if on_track else RED_RAG, align=CENTER,
        font=GREEN_FONT if on_track else RED_FONT)
    val(ws1, R, 6, "Higher ▲" if higher else "Lower ▼", fill=fill, align=CENTER)
    ws1.row_dimensions[R].height = 22
    R += 1

# Column widths
for ci, w in enumerate([28, 20, 20, 12, 16, 14], 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: DETAILED SUGGESTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Suggestions")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H1")
c = ws2["A1"]
c.value = "ACTIONABLE SUGGESTIONS — FUNNEL / PRICING / RULES / COLLECTION"
c.font, c.fill, c.alignment, c.border = TITLE_FONT, LIGHT_BLUE, CENTER, THIN_BORDER
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A2:H2")
c = ws2["A2"]
c.value = ("Review each suggestion below. Accepted changes should be applied to the corresponding "
           "Excel config file (Loan Pricing.xlsx / Loan Eligibility Rules.xlsx) and the pipeline re-run.")
c.font, c.fill, c.alignment, c.border = NOTE_FONT, WHITE, LEFT, THIN_BORDER

R = 4

# Group suggestions by lever
levers = ["Funnel / Application Intake", "Loan Pricing Matrix",
          "Rule Engine / Eligibility", "Collection Policy"]

for lever in levers:
    lever_suggestions = [s for s in suggestions if s["Lever"] == lever]
    if not lever_suggestions:
        continue

    section_title(ws2, R, f"  {lever.upper()}", col_end=8)
    R += 1

    s_headers = ["#", "KPI", "Gap", "Suggestion", "Priority", "Current Value",
                 "Suggested Change", "Action (User)"]
    for ci, h in enumerate(s_headers, 1):
        hdr(ws2, R, ci, h, fill=BLUE, font=SUB_FONT)
    ws2.row_dimensions[R].height = 24
    R += 1

    for idx, s in enumerate(lever_suggestions, 1):
        fill = ALT if idx % 2 == 1 else WHITE

        # Priority color
        pri = s["Priority"]
        pri_fill = RED_RAG if pri == "High" else (YELLOW_RAG if pri == "Medium" else LIGHT_GREEN)

        val(ws2, R, 1, idx, fill=fill, align=CENTER)
        val(ws2, R, 2, s["KPI"], fill=fill, font=BOLD_BODY)
        val(ws2, R, 3, s["Gap"], fill=fill)
        val(ws2, R, 4, s["Suggestion"], fill=fill)
        val(ws2, R, 5, pri, fill=pri_fill, align=CENTER, font=BOLD_BODY)
        val(ws2, R, 6, s["Current Value"], fill=fill)
        val(ws2, R, 7, s["Suggested Change"], fill=fill)

        # Action column — user fills this
        action_cell = ws2.cell(row=R, column=8, value="")
        action_cell.fill = AMBER
        action_cell.border = THIN_BORDER
        action_cell.alignment = WRAP

        ws2.row_dimensions[R].height = 80
        R += 1

    R += 1

# Column widths
for ci, w in enumerate([5, 18, 30, 55, 10, 30, 35, 20], 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: SEGMENT DATA ANALYSIS (supporting data for suggestions)
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Data Analysis")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:G1")
c = ws3["A1"]
c.value = "SEGMENT-LEVEL ANALYSIS — SUPPORTING DATA"
c.font, c.fill, c.alignment, c.border = TITLE_FONT, LIGHT_BLUE, CENTER, THIN_BORDER
ws3.row_dimensions[1].height = 36

ws3.merge_cells("A2:G2")
c = ws3["A2"]
c.value = "Drill-down data used to generate suggestions. Identifies which segments/ratings drive KPI gaps."
c.font, c.fill, c.alignment, c.border = NOTE_FONT, WHITE, LEFT, THIN_BORDER

R = 4
section_title(ws3, R, "  NPA ANALYSIS BY SEGMENT", col_end=7)
R += 1
for ci, h in enumerate(["Segment", "Loan Count", "NPA Count", "NPA %",
                         "Outstanding (Rs)", "Avg Rate", "Yield"], 1):
    hdr(ws3, R, ci, h, fill=BLUE, font=SUB_FONT)
ws3.row_dimensions[R].height = 24
R += 1

for i, (_, row) in enumerate(seg_npa.iterrows()):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws3, R, 1, row["Segment"], fill=fill, font=BOLD_BODY)
    val(ws3, R, 2, int(row["Count"]), fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 3, int(row["NPA"]), fmt="0", fill=fill, align=RIGHT)
    npa_fill = RED_RAG if row["NPA%"] > 0.05 else (YELLOW_RAG if row["NPA%"] > 0.03 else fill)
    val(ws3, R, 4, row["NPA%"], fmt="0.00%", fill=npa_fill, align=RIGHT)
    val(ws3, R, 5, row["Outstanding"], fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 6, row["Avg_Rate"], fmt="0.00%", fill=fill, align=RIGHT)
    val(ws3, R, 7, row["Yield"], fmt="0.00%", fill=fill, align=RIGHT)
    ws3.row_dimensions[R].height = 20
    R += 1

R += 2
section_title(ws3, R, "  NPA ANALYSIS BY CREDIT RATING", col_end=7)
R += 1
for ci, h in enumerate(["Rating", "Loan Count", "NPA Count", "NPA %",
                         "Outstanding (Rs)", "Avg Rate", "Avg Monthly Profit"], 1):
    hdr(ws3, R, ci, h, fill=BLUE, font=SUB_FONT)
ws3.row_dimensions[R].height = 24
R += 1

for i, (_, row) in enumerate(rat_npa.iterrows()):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws3, R, 1, row["Credit_Rating"], fill=fill, font=BOLD_BODY)
    val(ws3, R, 2, int(row["Count"]), fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 3, int(row["NPA"]), fmt="0", fill=fill, align=RIGHT)
    npa_fill = RED_RAG if row["NPA%"] > 0.10 else (YELLOW_RAG if row["NPA%"] > 0.05 else fill)
    val(ws3, R, 4, row["NPA%"], fmt="0.00%", fill=npa_fill, align=RIGHT)
    val(ws3, R, 5, row["Outstanding"], fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 6, row["Avg_Rate"], fmt="0.00%", fill=fill, align=RIGHT)
    val(ws3, R, 7, row["Avg_Profit"], fmt="#,##0.00", fill=fill, align=RIGHT)
    ws3.row_dimensions[R].height = 20
    R += 1

R += 2
section_title(ws3, R, "  NPA ANALYSIS BY PRODUCT", col_end=7)
R += 1
for ci, h in enumerate(["Product", "Loan Count", "NPA Count", "NPA %",
                         "Outstanding (Rs)", "Yield", "—"], 1):
    hdr(ws3, R, ci, h, fill=BLUE, font=SUB_FONT)
ws3.row_dimensions[R].height = 24
R += 1

for i, (_, row) in enumerate(prod_npa.iterrows()):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws3, R, 1, row["Product"], fill=fill, font=BOLD_BODY)
    val(ws3, R, 2, int(row["Count"]), fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 3, int(row["NPA"]), fmt="0", fill=fill, align=RIGHT)
    npa_fill = RED_RAG if row["NPA%"] > 0.10 else (YELLOW_RAG if row["NPA%"] > 0.05 else fill)
    val(ws3, R, 4, row["NPA%"], fmt="0.00%", fill=npa_fill, align=RIGHT)
    val(ws3, R, 5, row["Outstanding"], fmt="#,##0", fill=fill, align=RIGHT)
    val(ws3, R, 6, row["Yield"], fmt="0.00%", fill=fill, align=RIGHT)
    val(ws3, R, 7, "", fill=fill)
    ws3.row_dimensions[R].height = 20
    R += 1

# Column widths for data analysis
for ci, w in enumerate([18, 14, 12, 10, 22, 12, 18], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "KPI_Gap_Suggestions.xlsx")
wb.save(out_path)
print(f"\nDone! KPI Gap Suggestions saved to: {out_path}")
print(f"  - Sheet 1: Gap Summary       ({len(kpi_names)} KPIs compared)")
print(f"  - Sheet 2: Suggestions       ({len(suggestions)} actionable items)")
print(f"  - Sheet 3: Data Analysis      (segment/rating/product drill-down)")
