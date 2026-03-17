"""
Repayment Assumptions Generator
Generates an editable Excel file with all repayment simulation assumptions.
User should review/adjust before running generate_repayment_schedule.py.

Sections:
  1. Payment Profile Distribution   (% of loans per behaviour profile)
  2. DPD Distribution — Occasional Delay
  3. DPD Distribution — Chronic Delay
  4. Early (Partial) Payment parameters
  5. Full Loan Payment / Prepayment parameters
  6. Write-Offs & Penal Rules

Usage:
    python build_repayment_assumptions.py
    python build_repayment_assumptions.py --output "C:/path/to/folder"
"""

import argparse
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--output", default=r"C:\Users\joshm\OneDrive\Documents\BRE")
args = parser.parse_args()
OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY        = PatternFill("solid", fgColor="1F4E79")
BLUE        = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F0")
YELLOW      = PatternFill("solid", fgColor="FFF2CC")   # editable
GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")
WHITE       = PatternFill("solid", fgColor="FFFFFF")
ORANGE_LIGHT= PatternFill("solid", fgColor="FCE4D6")   # totals / warning

THIN_SIDE   = Side(style="thin",   color="AAAAAA")
MED_SIDE    = Side(style="medium", color="1F4E79")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
SEC_FONT   = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
VAL_FONT   = Font(bold=True, color="7B3F00", size=10)
NOTE_FONT  = Font(italic=True, color="595959", size=9)
TOTAL_FONT = Font(bold=True, color="9C0006", size=10)


def _c(ws, row, col, value, fill=WHITE, font=None, align=LEFT, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.alignment, c.border = fill, align, THIN_BORDER
    if font: c.font = font
    if fmt:  c.number_format = fmt
    return c


def sec_hdr(ws, row, title, ncols=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.fill, c.font, c.alignment, c.border = NAVY, SEC_FONT, LEFT, THIN_BORDER
    ws.row_dimensions[row].height = 22
    return row + 1


def col_hdrs(ws, row, headers, fills=None):
    for ci, h in enumerate(headers, 1):
        f = (fills[ci-1] if fills else BLUE)
        _c(ws, row, ci, h, fill=f, font=HDR_FONT, align=CENTER)
    ws.row_dimensions[row].height = 26
    return row + 1


# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Repayment Assumptions"
ws.sheet_view.showGridLines = False

col_widths = {1: 38, 2: 18, 3: 28, 4: 28}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:D1")
c = ws["A1"]
c.value     = "REPAYMENT SCHEDULE — SIMULATION ASSUMPTIONS"
c.font      = TITLE_FONT
c.fill      = LIGHT_BLUE
c.alignment = CENTER
c.border    = THIN_BORDER
ws.row_dimensions[1].height = 34

ws.merge_cells("A2:D2")
c = ws["A2"]
c.value = (f"Generated: {datetime.today().strftime('%d %b %Y')}  |  "
           "Edit amber cells to match your portfolio assumptions.  "
           "Percentage columns must sum to 100%.  "
           "Save & close before running generate_repayment_schedule.py.")
c.font      = NOTE_FONT
c.fill      = WHITE
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border    = THIN_BORDER
ws.row_dimensions[2].height = 28

R = 4

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Payment Profile Distribution
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 1 — PAYMENT PROFILE DISTRIBUTION  (must sum to 100%)")
R = col_hdrs(ws, R, ["Profile", "% of Loans ✏", "Description", ""])

PROFILES = [
    ("Pristine",          45, "Always pays on time. DPD = 0 every month."),
    ("Occasional Delay",  25, "Mostly on time, but misses deadline occasionally (see Section 2)."),
    ("Chronic Delay",     10, "Frequently late. DPD pattern per Section 3."),
    ("Partial Prepayment",10, "On time, makes one partial prepayment (see Section 4)."),
    ("Full Prepayment",   10, "On time, closes loan early in full (see Section 5)."),
]
p1_start = R
for i, (name, pct, desc) in enumerate(PROFILES):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, name, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, pct,  fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 3, desc, fill=fill, font=NOTE_FONT)
    _c(ws, R, 4, "",   fill=fill)
    R += 1
# Total row
_c(ws, R, 1, "TOTAL", fill=ORANGE_LIGHT, font=TOTAL_FONT)
_c(ws, R, 2, f"=SUM(B{p1_start}:B{R-1})", fill=ORANGE_LIGHT, font=TOTAL_FONT, align=RIGHT, fmt="0.00")
_c(ws, R, 3, "← Must equal 100", fill=ORANGE_LIGHT, font=NOTE_FONT)
_c(ws, R, 4, "", fill=ORANGE_LIGHT)
R += 2

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — DPD Distribution: Occasional Delay
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 2 — DPD DISTRIBUTION: OCCASIONAL DELAY  (must sum to 100%)")
R = col_hdrs(ws, R, ["DPD Scenario", "DPD Days ✏", "% Probability ✏", "Description"])

OCC_DPD = [
    ("On Time",      0,  70, "Payment made on due date"),
    ("Short Delay", 15,  20, "Payment 1–15 days late"),
    ("Medium Delay",45,  10, "Payment 31–45 days late"),
]
s2_start = R
for i, (label, days, prob, desc) in enumerate(OCC_DPD):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, label, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, days,  fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0")
    _c(ws, R, 3, prob,  fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 4, desc,  fill=fill, font=NOTE_FONT)
    R += 1
_c(ws, R, 1, "TOTAL", fill=ORANGE_LIGHT, font=TOTAL_FONT)
_c(ws, R, 2, "",      fill=ORANGE_LIGHT)
_c(ws, R, 3, f"=SUM(C{s2_start}:C{R-1})", fill=ORANGE_LIGHT, font=TOTAL_FONT, align=RIGHT, fmt="0.00")
_c(ws, R, 4, "← Must equal 100", fill=ORANGE_LIGHT, font=NOTE_FONT)
R += 2

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — DPD Distribution: Chronic Delay
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 3 — DPD DISTRIBUTION: CHRONIC DELAY  (must sum to 100%)")
R = col_hdrs(ws, R, ["DPD Scenario", "DPD Days ✏", "% Probability ✏", "Description"])

CHR_DPD = [
    ("On Time",       0, 30, "Occasional on-time payment"),
    ("Short Delay",  20, 30, "Payment 16–20 days late"),
    ("Medium Delay", 45, 20, "Payment 31–45 days late"),
    ("Long Delay",   75, 15, "Payment 61–75 days late — approaching NPA"),
    ("NPA Risk",     95,  5, "Payment 91–95 days late — NPA"),
]
s3_start = R
for i, (label, days, prob, desc) in enumerate(CHR_DPD):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, label, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, days,  fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0")
    _c(ws, R, 3, prob,  fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 4, desc,  fill=fill, font=NOTE_FONT)
    R += 1
_c(ws, R, 1, "TOTAL", fill=ORANGE_LIGHT, font=TOTAL_FONT)
_c(ws, R, 2, "",      fill=ORANGE_LIGHT)
_c(ws, R, 3, f"=SUM(C{s3_start}:C{R-1})", fill=ORANGE_LIGHT, font=TOTAL_FONT, align=RIGHT, fmt="0.00")
_c(ws, R, 4, "← Must equal 100", fill=ORANGE_LIGHT, font=NOTE_FONT)
R += 2

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — Early (Partial) Payment
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 4 — EARLY (PARTIAL) PAYMENT ASSUMPTIONS")
R = col_hdrs(ws, R, ["Parameter", "Value ✏", "Unit", "Description"])

PARTIAL_PARAMS = [
    ("Min partial prepayment % of outstanding",  10,  "%",     "Minimum amount prepaid as % of remaining balance"),
    ("Max partial prepayment % of outstanding",  30,  "%",     "Maximum amount prepaid as % of remaining balance"),
    ("Earliest EMI month for partial prepayment", 3,  "Month", "Cannot prepay before this EMI number"),
    ("Prepayment charge on partial amount",        2,  "%",     "Fee charged as % of the prepaid amount"),
]
for i, (param, val, unit, desc) in enumerate(PARTIAL_PARAMS):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, param, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, val,   fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 3, unit,  fill=fill, font=NOTE_FONT, align=CENTER)
    _c(ws, R, 4, desc,  fill=fill, font=NOTE_FONT)
    R += 1
R += 1

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — Full Loan Payment / Prepayment
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 5 — FULL LOAN PAYMENT (EARLY CLOSURE)")
R = col_hdrs(ws, R, ["Parameter", "Value ✏", "Unit", "Description"])

FULL_PARAMS = [
    ("Earliest EMI month for full closure",  3,  "Month", "Loan cannot be closed before this EMI number"),
    ("Latest EMI month for full closure",   36,  "Month", "Random closure chosen between earliest and this month"),
    ("Prepayment charge on outstanding",     2,  "%",     "Fee charged as % of remaining outstanding principal"),
]
for i, (param, val, unit, desc) in enumerate(FULL_PARAMS):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, param, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, val,   fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 3, unit,  fill=fill, font=NOTE_FONT, align=CENTER)
    _c(ws, R, 4, desc,  fill=fill, font=NOTE_FONT)
    R += 1
R += 1

# ══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — Write-Offs & Penal Rules
# ══════════════════════════════════════════════════════════════════════════════
R = sec_hdr(ws, R, "  SECTION 6 — WRITE-OFFS & PENAL RULES")
R = col_hdrs(ws, R, ["Parameter", "Value ✏", "Unit", "Description"])

PENAL_PARAMS = [
    ("NPA classification DPD threshold",         90,  "Days",   "Loan classified as NPA when DPD exceeds this value"),
    ("Default penalty (flat fee) when NPA",    1000,  "Rs",     "One-time flat penalty charged per EMI when DPD > NPA threshold"),
    ("Penal interest rate (p.a.)",                2,  "% p.a.", "Annual rate applied daily on OS principal when DPD > 0"),
    ("Write-off DPD threshold",                 180,  "Days",   "Loan written off when DPD exceeds this value"),
    ("Write-off % of outstanding balance",       30,  "%",      "Portion of outstanding principal written off at threshold"),
]
for i, (param, val, unit, desc) in enumerate(PENAL_PARAMS):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    _c(ws, R, 1, param, fill=fill, font=LABEL_FONT)
    _c(ws, R, 2, val,   fill=YELLOW, font=VAL_FONT, align=RIGHT, fmt="0.00")
    _c(ws, R, 3, unit,  fill=fill, font=NOTE_FONT, align=CENTER)
    _c(ws, R, 4, desc,  fill=fill, font=NOTE_FONT)
    R += 1

ws.freeze_panes = "A3"

# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "Repayment Assumptions.xlsx")
wb.save(out_path)
print(f"Repayment Assumptions saved: {out_path}")
print(f"  Sections: 1-Payment Profiles  2-DPD Occasional  3-DPD Chronic")
print(f"            4-Partial Prepay    5-Full Prepay     6-Write-offs & Penal")
print()
print("  NEXT STEP: Open the file, review/adjust amber cells, ensure")
print("  % columns in Sections 1-3 each sum to 100, save and close.")
