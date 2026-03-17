"""
Loan Repayment Schedule Generator
Reads approved_loans.xlsx and Repayment Assumptions.xlsx, generates monthly
repayment schedules applying configurable BRE rules and payment profiles.

All simulation parameters (profile weights, DPD distributions, prepayment
parameters, penal rates, write-off thresholds) are loaded from
Repayment Assumptions.xlsx (built by build_repayment_assumptions.py).
Falls back to built-in defaults if the file is missing.

Usage: python generate_repayment_schedule.py
       python generate_repayment_schedule.py
           --input       "path/to/approved_loans.xlsx"
           --assumptions "path/to/Repayment Assumptions.xlsx"
           --output      "path/to/output"
"""

import argparse
import os
import random
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--input",       default=r"C:\Users\joshm\OneDrive\Documents\BRE\approved_loans.xlsx")
parser.add_argument("--assumptions", default=r"C:\Users\joshm\OneDrive\Documents\BRE\Repayment Assumptions.xlsx")
parser.add_argument("--output",      default=r"C:\Users\joshm\OneDrive\Documents\BRE")
args = parser.parse_args()

OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)
random.seed(42)
np.random.seed(42)

# ── 1. LOAD APPROVED LOANS ────────────────────────────────────────────────────
df = pd.read_excel(args.input, sheet_name="Approved_Loans")
print(f"Loaded {len(df):,} approved loans")

# ── 2. LOAD REPAYMENT ASSUMPTIONS ─────────────────────────────────────────────
# Built-in defaults
_D = {
    # Section 1 — profile weights
    "profile_weights": {"Pristine": 45, "Occasional Delay": 25, "Chronic Delay": 10,
                        "Partial Prepay": 10, "Full Prepay": 10},
    # Section 2 — occasional delay DPD
    "occ_dpd":   [0, 15, 45],
    "occ_probs": [70, 20, 10],
    # Section 3 — chronic delay DPD
    "chr_dpd":   [0, 20, 45, 75, 95],
    "chr_probs": [30, 30, 20, 15, 5],
    # Section 4 — partial prepayment
    "partial_min_pct":   10,   # % of outstanding
    "partial_max_pct":   30,
    "partial_min_month":  3,
    "partial_charge_pct": 2,   # % of prepaid amount
    # Section 5 — full prepayment
    "full_min_month":   3,
    "full_max_month":  36,
    "full_charge_pct":  2,     # % of outstanding
    # Section 6 — penal rules
    "npa_dpd_threshold":  90,
    "default_penalty":  1000,
    "penal_rate_pa":       2,  # % p.a.
    "writeoff_dpd":      180,
    "writeoff_pct":       30,  # % of outstanding
}

A = dict(_D)  # will be overridden from file

if os.path.exists(args.assumptions):
    try:
        xl = pd.read_excel(args.assumptions, sheet_name="Repayment Assumptions", header=None)
        rows = xl.values.tolist()

        def _find_section(keyword):
            """Return row index of header row after the section title containing keyword."""
            for i, row in enumerate(rows):
                if any(isinstance(v, str) and keyword.lower() in str(v).lower() for v in row):
                    return i
            return None

        import math as _math

        def _is_real(v):
            """True if v is a non-NaN, non-None value."""
            if v is None: return False
            try:
                return not _math.isnan(float(v))
            except (TypeError, ValueError):
                return isinstance(v, str) and v.strip() != ""

        _SKIP = {"", "nan", "none", "total", "parameter", "profile",
                 "dpd scenario", "label", "description", "unit"}

        def _valid_label(v):
            return _is_real(v) and str(v).strip().lower() not in _SKIP

        def _to_f(v):
            try:
                fv = float(v)
                return None if _math.isnan(fv) else fv
            except: return None

        def _read_table(start_row, col_label=0, col_val=1, max_rows=10):
            """Read label→float value pairs, skipping blank/header/total rows."""
            result = {}
            for r in rows[start_row+1 : start_row+1+max_rows]:
                if max(col_label, col_val) >= len(r): continue
                if not _valid_label(r[col_label]): continue
                fv = _to_f(r[col_val])
                if fv is not None:
                    result[str(r[col_label]).strip()] = fv
            return result

        def _read_dpd_table(start_row, max_rows):
            """Read (days, prob) pairs from cols 1 and 2 of a DPD section."""
            day_list, prob_list = [], []
            for r in rows[start_row+1 : start_row+1+max_rows]:
                if len(r) < 3: continue
                if not _valid_label(r[0]): continue
                d = _to_f(r[1]);  p = _to_f(r[2])
                if d is not None and p is not None:
                    day_list.append(int(d));  prob_list.append(p)
            return day_list, prob_list

        # Section 1 — profile weights
        s1 = _find_section("SECTION 1")
        if s1 is not None:
            t = _read_table(s1 + 1, col_label=0, col_val=1, max_rows=6)
            for k, v in t.items():
                if "Pristine" in k:     A["profile_weights"]["Pristine"]         = v
                elif "Occasional" in k: A["profile_weights"]["Occasional Delay"] = v
                elif "Chronic" in k:    A["profile_weights"]["Chronic Delay"]    = v
                elif "Partial" in k:    A["profile_weights"]["Partial Prepay"]   = v
                elif "Full" in k:       A["profile_weights"]["Full Prepay"]      = v

        # Section 2 — occasional delay
        s2 = _find_section("SECTION 2")
        if s2 is not None:
            days_list, prob_list = _read_dpd_table(s2 + 1, max_rows=5)
            if days_list and prob_list and len(days_list) == len(prob_list):
                A["occ_dpd"]   = days_list
                A["occ_probs"] = prob_list

        # Section 3 — chronic delay
        s3 = _find_section("SECTION 3")
        if s3 is not None:
            days_list, prob_list = _read_dpd_table(s3 + 1, max_rows=7)
            if days_list and prob_list and len(days_list) == len(prob_list):
                A["chr_dpd"]   = days_list
                A["chr_probs"] = prob_list

        # Section 4 — partial prepayment
        s4 = _find_section("SECTION 4")
        if s4 is not None:
            t = _read_table(s4 + 1, col_label=0, col_val=1, max_rows=6)
            for k, v in t.items():
                k_l = k.lower()
                if "min partial" in k_l or "min %" in k_l:    A["partial_min_pct"]   = v
                elif "max partial" in k_l or "max %" in k_l:  A["partial_max_pct"]   = v
                elif "earliest" in k_l:                        A["partial_min_month"] = int(v)
                elif "charge" in k_l:                          A["partial_charge_pct"]= v

        # Section 5 — full prepayment
        s5 = _find_section("SECTION 5")
        if s5 is not None:
            t = _read_table(s5 + 1, col_label=0, col_val=1, max_rows=5)
            for k, v in t.items():
                k_l = k.lower()
                if "earliest" in k_l:    A["full_min_month"]  = int(v)
                elif "latest" in k_l:    A["full_max_month"]  = int(v)
                elif "charge" in k_l:    A["full_charge_pct"] = v

        # Section 6 — penal rules
        s6 = _find_section("SECTION 6")
        if s6 is not None:
            t = _read_table(s6 + 1, col_label=0, col_val=1, max_rows=7)
            for k, v in t.items():
                k_l = k.lower()
                if "npa classification" in k_l:  A["npa_dpd_threshold"] = int(v)
                elif "default penalty" in k_l:   A["default_penalty"]   = float(v)
                elif "penal interest" in k_l:    A["penal_rate_pa"]     = v
                elif "write-off dpd" in k_l:     A["writeoff_dpd"]      = int(v)
                elif "write-off %" in k_l:        A["writeoff_pct"]      = v

        print(f"Repayment assumptions loaded from {args.assumptions}")
    except Exception as e:
        print(f"WARNING: Could not read assumptions ({e}) — using built-in defaults")
else:
    print(f"WARNING: Assumptions file not found at {args.assumptions} — using built-in defaults")

# Print active assumptions
pw = A["profile_weights"]
print(f"\n  Active assumptions:")
print(f"    Profiles  : Pristine {pw['Pristine']}%  OccDelay {pw['Occasional Delay']}%  "
      f"Chronic {pw['Chronic Delay']}%  PartialPrepay {pw['Partial Prepay']}%  FullPrepay {pw['Full Prepay']}%")
print(f"    NPA DPD   : >{A['npa_dpd_threshold']} days  |  Penal Rate: {A['penal_rate_pa']}% p.a.")
print(f"    Default Penalty: Rs {A['default_penalty']:,.0f}  |  "
      f"Write-off at DPD>{A['writeoff_dpd']} ({A['writeoff_pct']}% of OS)")
print(f"    Partial Prepay: {A['partial_min_pct']}–{A['partial_max_pct']}% of OS  "
      f"from month {A['partial_min_month']}")
print(f"    Full Prepay   : month {A['full_min_month']}–{A['full_max_month']}  "
      f"charge {A['full_charge_pct']}%")
print()

# ── 3. BUILD PROFILES FROM ASSUMPTIONS ────────────────────────────────────────
_occ_probs_norm = [p / sum(A["occ_probs"]) for p in A["occ_probs"]]
_chr_probs_norm = [p / sum(A["chr_probs"]) for p in A["chr_probs"]]

PROFILES = {
    "Pristine":         {"weight": pw["Pristine"],
                         "dpd_choices": [0], "dpd_probs": [1.0]},
    "Occasional Delay": {"weight": pw["Occasional Delay"],
                         "dpd_choices": A["occ_dpd"], "dpd_probs": _occ_probs_norm},
    "Chronic Delay":    {"weight": pw["Chronic Delay"],
                         "dpd_choices": A["chr_dpd"], "dpd_probs": _chr_probs_norm},
    "Partial Prepay":   {"weight": pw["Partial Prepay"],
                         "dpd_choices": [0], "dpd_probs": [1.0], "has_partial_prepay": True},
    "Full Prepay":      {"weight": pw["Full Prepay"],
                         "dpd_choices": [0], "dpd_probs": [1.0], "has_full_prepay": True},
}

profile_names   = list(PROFILES.keys())
profile_weights = [PROFILES[p]["weight"] for p in profile_names]
loan_profiles   = random.choices(profile_names, weights=profile_weights, k=len(df))

# ── 4. BRE RULE FUNCTIONS (parameterised from assumptions) ────────────────────
_PENAL_RATE   = A["penal_rate_pa"] / 100
_NPA_THRESH   = A["npa_dpd_threshold"]
_DEF_PENALTY  = A["default_penalty"]
_FULL_CHG_PCT = A["full_charge_pct"] / 100
_PART_CHG_PCT = A["partial_charge_pct"] / 100

def bre_penal_interest(dpd, outstanding_principal):
    """Penal interest at configured rate (p.a.) applied daily on OS principal when DPD > 0."""
    if dpd > 0:
        return round(outstanding_principal * (_PENAL_RATE / 365) * dpd, 2)
    return 0.0

def bre_dpd_bucket(dpd):
    if dpd == 0:    return "Current"
    elif dpd <= 30: return "DPD 1-30"
    elif dpd <= 60: return "DPD 31-60"
    elif dpd <= 90: return "DPD 61-90"
    else:           return "DPD 90+"

def bre_default_penalty(dpd):
    return _DEF_PENALTY if dpd > _NPA_THRESH else 0

def bre_loan_status(dpd):
    return "NPA" if dpd > _NPA_THRESH else "STANDARD"

def bre_prepayment_charge(prepayment_amount, outstanding_principal, is_full):
    rate = _FULL_CHG_PCT if is_full else _PART_CHG_PCT
    base = outstanding_principal if is_full else prepayment_amount
    return round(base * rate, 2)

# ── 4. GENERATE SCHEDULE ──────────────────────────────────────────────────────
schedule_rows = []
sched_id      = 1

for idx, loan in df.iterrows():
    profile_name = loan_profiles[idx]
    profile      = PROFILES[profile_name]

    principal      = float(loan["Sanctioned_Amount"])
    annual_rate    = float(loan["Interest_Rate"])
    tenure         = int(loan["Loan_Tenure_Months"])
    scheduled_emi  = float(loan["EMI"])
    loan_id        = loan["Loan_ID"]
    monthly_rate   = annual_rate / 12
    disbursement_dt = datetime.strptime(str(loan["Disbursement_Date"])[:10], "%Y-%m-%d")

    # Assign prepayment month from assumptions
    p_min  = min(A["partial_min_month"], max(3, tenure - 1))
    p_safe = list(range(p_min, max(p_min + 1, tenure)))
    partial_prepay_month = random.choice(p_safe) if profile.get("has_partial_prepay") else None

    f_min  = min(A["full_min_month"], max(3, tenure - 1))
    f_max  = min(A["full_max_month"], max(f_min + 1, tenure - 1))
    f_safe = list(range(f_min, max(f_min + 1, f_max + 1)))
    full_prepay_month    = random.choice(f_safe) if profile.get("has_full_prepay") else None

    opening_balance = principal

    for emi_num in range(1, tenure + 1):
        due_date = disbursement_dt + timedelta(days=30 * emi_num)

        # Amortization components
        interest_component  = round(opening_balance * monthly_rate, 2)
        regular_principal   = round(min(max(0, scheduled_emi - interest_component),
                                        opening_balance), 2)

        # DPD for this EMI (drawn from profile distribution)
        dpd = random.choices(profile["dpd_choices"], weights=profile["dpd_probs"])[0]
        payment_dt = due_date + timedelta(days=dpd)

        # ── BRE RS2 ──────────────────────────────────────────────────────────
        penal_interest = bre_penal_interest(dpd, opening_balance)
        dpd_bucket     = bre_dpd_bucket(dpd)

        # ── BRE RS3 ──────────────────────────────────────────────────────────
        default_penalty = bre_default_penalty(dpd)
        loan_status     = bre_loan_status(dpd)

        # ── BRE RS4 — Prepayment ─────────────────────────────────────────────
        prepayment_amount  = 0.0
        prepayment_charge  = 0.0
        prepayment_type    = ""
        is_full_prepay     = False

        if emi_num == full_prepay_month:
            # Full prepayment: pay off remaining balance above regular EMI principal
            prepayment_amount = round(max(0, opening_balance - regular_principal), 2)
            prepayment_charge = bre_prepayment_charge(0, opening_balance, is_full=True)
            prepayment_type   = "FULL_PREPAYMENT"
            is_full_prepay    = True

        elif emi_num == partial_prepay_month:
            # Partial prepayment: configured % range of current outstanding
            partial_pct       = np.random.uniform(A["partial_min_pct"] / 100,
                                                   A["partial_max_pct"] / 100)
            prepayment_amount = round(opening_balance * partial_pct, 2)
            prepayment_charge = bre_prepayment_charge(prepayment_amount, 0, is_full=False)
            prepayment_type   = "PARTIAL_PREPAYMENT"

        # Closing balance
        closing_balance = round(max(0, opening_balance - regular_principal - prepayment_amount), 2)

        # Payment status
        if is_full_prepay:
            payment_status = "Full Prepayment"
        elif prepayment_amount > 0:
            payment_status = "Partial Prepayment"
        elif dpd > 90:
            payment_status = "Defaulted"
        elif dpd > 0:
            payment_status = "Delayed"
        else:
            payment_status = "On Time"

        # Totals
        total_due  = round(scheduled_emi + penal_interest + default_penalty, 2)
        total_paid = round(scheduled_emi + penal_interest + default_penalty
                           + prepayment_amount + prepayment_charge, 2)

        schedule_rows.append([
            f"SCH-{sched_id:05d}",
            loan_id,
            emi_num,
            due_date.strftime("%Y-%m-%d"),
            payment_dt.strftime("%Y-%m-%d"),
            dpd,
            dpd_bucket,
            round(opening_balance, 2),
            scheduled_emi,
            regular_principal,
            interest_component,
            penal_interest,
            prepayment_amount,
            prepayment_charge,
            prepayment_type,
            default_penalty,
            loan_status,
            closing_balance,
            payment_status,
            total_due,
            total_paid,
        ])

        sched_id      += 1
        opening_balance = closing_balance

        if is_full_prepay or closing_balance <= 0.01:
            break  # Loan fully repaid

SCHED_COLS = [
    "Schedule_ID", "Loan_ID", "EMI_Number", "Due_Date", "Payment_Date",
    "DPD", "DPD_Bucket", "Opening_Balance", "Scheduled_EMI",
    "Principal_Component", "Interest_Component", "Penal_Interest",
    "Prepayment_Amount", "Prepayment_Charge", "Prepayment_Type",
    "Default_Penalty", "Loan_Status", "Closing_Balance",
    "Payment_Status", "Total_Amount_Due", "Total_Amount_Paid",
]
sched_df = pd.DataFrame(schedule_rows, columns=SCHED_COLS)
NUM_ROWS = len(sched_df)
print(f"Schedule rows generated: {NUM_ROWS:,}")

# ── 5. STYLES ─────────────────────────────────────────────────────────────────
NAVY       = PatternFill("solid", fgColor="1F4E79")
BLUE       = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE = PatternFill("solid", fgColor="D6E4F0")
ALT        = PatternFill("solid", fgColor="EBF5FB")
WHITE      = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, color="1F4E79", size=16)
LABEL_FONT = Font(bold=True, color="1F4E79", size=10)
KPI_FONT   = Font(bold=True, color="1F4E79", size=12)
SUB_FONT   = Font(bold=True, color="FFFFFF", size=10)

THIN_SIDE   = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def hdr(ws, row, col, value, fill=NAVY, font=HDR_FONT, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = fill, font, align, THIN_BORDER
    return c


def val(ws, row, col, value, fmt=None, fill=WHITE, align=RIGHT):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.fill, c.alignment, c.border = fill, align, THIN_BORDER
    return c


def section_title(ws, row, title, col_start=1, col_end=6):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.fill      = NAVY
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = THIN_BORDER
    ws.row_dimensions[row].height = 22


# ── 6. WORKBOOK ───────────────────────────────────────────────────────────────
wb = Workbook()

# ── REPAYMENT SCHEDULE SHEET ──────────────────────────────────────────────────
ws_s = wb.active
ws_s.title = "Repayment_Schedule"

for ci, name in enumerate(SCHED_COLS, 1):
    hdr(ws_s, 1, ci, name)

# Col index → number format
col_fmts = {
    8:  "#,##0.00",   # Opening_Balance
    9:  "#,##0.00",   # Scheduled_EMI
    10: "#,##0.00",   # Principal_Component
    11: "#,##0.00",   # Interest_Component
    12: "#,##0.00",   # Penal_Interest
    13: "#,##0.00",   # Prepayment_Amount
    14: "#,##0.00",   # Prepayment_Charge
    16: "#,##0.00",   # Default_Penalty
    18: "#,##0.00",   # Closing_Balance
    20: "#,##0.00",   # Total_Amount_Due
    21: "#,##0.00",   # Total_Amount_Paid
}

for ri, row_data in enumerate(sched_df.itertuples(index=False), 2):
    for ci, v in enumerate(row_data, 1):
        c = ws_s.cell(row=ri, column=ci, value=v)
        c.border = THIN_BORDER
        if ci in col_fmts:
            c.number_format = col_fmts[ci]

col_widths_s = [12, 14, 10, 12, 12, 6, 12, 16, 14, 18, 18, 14,
                18, 16, 18, 14, 12, 16, 16, 16, 16]
for i, w in enumerate(col_widths_s, 1):
    ws_s.column_dimensions[get_column_letter(i)].width = w
ws_s.freeze_panes = "A2"

# ── SUMMARY SHEET ─────────────────────────────────────────────────────────────
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False

for col, w in {1: 28, 2: 18, 3: 18, 4: 18, 5: 20, 6: 18}.items():
    ws.column_dimensions[get_column_letter(col)].width = w

last_row = NUM_ROWS + 1

def rng(col):
    return f"Repayment_Schedule!${col}$2:${col}${last_row}"

DPD_R    = rng("F")   # DPD
BUCKET   = rng("G")   # DPD_Bucket
SCH_EMI  = rng("I")   # Scheduled_EMI
PRIN     = rng("J")   # Principal_Component
INT_C    = rng("K")   # Interest_Component
PENAL    = rng("L")   # Penal_Interest
PREPAY   = rng("M")   # Prepayment_Amount
PREP_CHG = rng("N")   # Prepayment_Charge
DEF_PEN  = rng("P")   # Default_Penalty
LSTATUS  = rng("Q")   # Loan_Status
CLBAL    = rng("R")   # Closing_Balance
PSTATUS  = rng("S")   # Payment_Status
T_DUE    = rng("T")   # Total_Amount_Due
T_PAID   = rng("U")   # Total_Amount_Paid
SCHED_ID = rng("A")   # Schedule_ID (for COUNTA)

today = datetime(2026, 3, 16)

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value     = "LOAN REPAYMENT SCHEDULE — SUMMARY"
c.font      = TITLE_FONT
c.fill      = LIGHT_BLUE
c.alignment = CENTER
c.border    = THIN_BORDER
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
c = ws["A2"]
c.value = (f"Generated: {today.strftime('%d %b %Y')} | "
           f"Schedule Rows: {NUM_ROWS:,} | "
           f"BRE Rules: Penal Interest 2% p.a. | Default Penalty Rs 1,000 | Prepayment 2%")
c.font      = Font(italic=True, color="666666", size=9)
c.fill      = WHITE
c.alignment = CENTER
c.border    = THIN_BORDER

R = 4

# ── SECTION 1: Executive KPIs ─────────────────────────────────────────────────
section_title(ws, R, "  EXECUTIVE KPIs")
R += 1

kpis = [
    ("Total EMI Rows",              f"=COUNTA({SCHED_ID})",                                          "0"),
    ("Total Scheduled EMI (Rs)",    f"=SUM({SCH_EMI})",                                              "#,##0.00"),
    ("Total Interest Collected (Rs)",f"=SUM({INT_C})",                                               "#,##0.00"),
    ("Total Penal Interest (Rs)",   f"=SUM({PENAL})",                                                "#,##0.00"),
    ("Total Default Penalties (Rs)",f"=SUM({DEF_PEN})",                                              "#,##0.00"),
    ("Total Prepayments (Rs)",      f"=SUM({PREPAY})",                                               "#,##0.00"),
    ("On-Time Payment Rate (%)",    f'=IFERROR(COUNTIF({PSTATUS},"On Time")/COUNTA({SCHED_ID}),0)',  "0.00%"),
    ("Delayed Payment Count",       f'=COUNTIF({PSTATUS},"Delayed")',                                "0"),
    ("NPA EMI Rows",                f'=COUNTIF({LSTATUS},"NPA")',                                    "0"),
    ("Full Prepayment Count",       f'=COUNTIF({PSTATUS},"Full Prepayment")',                        "0"),
]

for idx, (label, formula, fmt) in enumerate(kpis):
    row = R + (idx // 2)
    col = 1 if idx % 2 == 0 else 4

    lc = ws.cell(row=row, column=col, value=label)
    lc.font, lc.fill, lc.alignment, lc.border = LABEL_FONT, LIGHT_BLUE, LEFT, THIN_BORDER

    ws.merge_cells(start_row=row, start_column=col+1,
                   end_row=row,   end_column=col+2)
    vc = ws.cell(row=row, column=col+1, value=formula)
    vc.font, vc.fill, vc.number_format = KPI_FONT, WHITE, fmt
    vc.alignment, vc.border = CENTER, THIN_BORDER

R += (len(kpis) + 1) // 2 + 1

# ── SECTION 2: Payment Status Distribution ───────────────────────────────────
section_title(ws, R, "  PAYMENT STATUS DISTRIBUTION")
R += 1

for ci, h in enumerate(["Payment Status", "Count", "% of Total",
                         "Total EMI (Rs)", "Total Penal Int (Rs)", "Total Penalties (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

statuses = ["On Time", "Delayed", "Defaulted", "Partial Prepayment", "Full Prepayment"]
for i, status in enumerate(statuses):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, status,                                                                   fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({PSTATUS},"{status}")',                       "#,##0",        fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({PSTATUS},"{status}")/COUNTA({SCHED_ID}),0)', "0.00%", fill=fill)
    val(ws, R, 4, f'=SUMIF({PSTATUS},"{status}",{SCH_EMI})',               "#,##0.00",     fill=fill)
    val(ws, R, 5, f'=SUMIF({PSTATUS},"{status}",{PENAL})',                 "#,##0.00",     fill=fill)
    val(ws, R, 6, f'=SUMIF({PSTATUS},"{status}",{DEF_PEN})',               "#,##0.00",     fill=fill)
    R += 1

R += 1

# ── SECTION 3: DPD Bucket Analysis ───────────────────────────────────────────
section_title(ws, R, "  DPD BUCKET ANALYSIS")
R += 1

for ci, h in enumerate(["DPD Bucket", "Count", "% of Total",
                         "Avg DPD", "Total Penal Interest (Rs)", "Total Penalties (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

dpd_buckets = ["Current", "DPD 1-30", "DPD 31-60", "DPD 61-90", "DPD 90+"]
for i, bucket in enumerate(dpd_buckets):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, bucket,                                                                    fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({BUCKET},"{bucket}")',                         "#,##0",        fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({BUCKET},"{bucket}")/COUNTA({SCHED_ID}),0)', "0.00%", fill=fill)
    val(ws, R, 4, f'=IFERROR(AVERAGEIF({BUCKET},"{bucket}",{DPD_R}),0)',    "0.0",          fill=fill)
    val(ws, R, 5, f'=SUMIF({BUCKET},"{bucket}",{PENAL})',                   "#,##0.00",     fill=fill)
    val(ws, R, 6, f'=SUMIF({BUCKET},"{bucket}",{DEF_PEN})',                 "#,##0.00",     fill=fill)
    R += 1

R += 1

# ── SECTION 4: Loan Status Summary ───────────────────────────────────────────
section_title(ws, R, "  LOAN STATUS SUMMARY (BRE RS3)")
R += 1

for ci, h in enumerate(["Loan Status", "EMI Row Count", "% of Total",
                         "Total Amount Due (Rs)", "Total Amount Paid (Rs)", "Total Penal Int (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, status in enumerate(["STANDARD", "NPA"]):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, status,                                                                    fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({LSTATUS},"{status}")',                        "#,##0",        fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({LSTATUS},"{status}")/COUNTA({SCHED_ID}),0)', "0.00%", fill=fill)
    val(ws, R, 4, f'=SUMIF({LSTATUS},"{status}",{T_DUE})',                  "#,##0.00",     fill=fill)
    val(ws, R, 5, f'=SUMIF({LSTATUS},"{status}",{T_PAID})',                 "#,##0.00",     fill=fill)
    val(ws, R, 6, f'=SUMIF({LSTATUS},"{status}",{PENAL})',                  "#,##0.00",     fill=fill)
    R += 1

R += 1

# ── SECTION 5: Prepayment Analysis ───────────────────────────────────────────
section_title(ws, R, "  PREPAYMENT ANALYSIS (BRE RS4)")
R += 1

for ci, h in enumerate(["Prepayment Type", "Count", "% of Total",
                         "Total Prepaid (Rs)", "Total Prepay Charge (Rs)", "Avg Prepaid Amount (Rs)"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1

for i, ptype in enumerate(["PARTIAL_PREPAYMENT", "FULL_PREPAYMENT"]):
    fill = ALT if i % 2 == 0 else WHITE
    PREPAY_TYPE = rng("O")   # Prepayment_Type column
    val(ws, R, 1, ptype,                                                                           fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({PREPAY_TYPE},"{ptype}")',                            "#,##0",        fill=fill)
    val(ws, R, 3, f'=IFERROR(COUNTIF({PREPAY_TYPE},"{ptype}")/COUNTA({SCHED_ID}),0)', "0.00%",    fill=fill)
    val(ws, R, 4, f'=SUMIF({PREPAY_TYPE},"{ptype}",{PREPAY})',                     "#,##0.00",     fill=fill)
    val(ws, R, 5, f'=SUMIF({PREPAY_TYPE},"{ptype}",{PREP_CHG})',                   "#,##0.00",     fill=fill)
    val(ws, R, 6, f'=IFERROR(AVERAGEIF({PREPAY_TYPE},"{ptype}",{PREPAY}),0)',      "#,##0.00",     fill=fill)
    R += 1

# ── 7. SAVE ───────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "loan_repayment_schedule.xlsx")
wb.save(out_path)
print(f"Repayment schedule saved: {out_path}")
print(f"  On Time     : {(sched_df['Payment_Status']=='On Time').sum():,}")
print(f"  Delayed     : {(sched_df['Payment_Status']=='Delayed').sum():,}")
print(f"  Defaulted   : {(sched_df['Payment_Status']=='Defaulted').sum():,}")
print(f"  Part Prepay : {(sched_df['Payment_Status']=='Partial Prepayment').sum():,}")
print(f"  Full Prepay : {(sched_df['Payment_Status']=='Full Prepayment').sum():,}")
