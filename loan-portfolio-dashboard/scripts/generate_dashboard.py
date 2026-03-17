"""
Loan Portfolio Dashboard Generator — Comprehensive Final Output
Generates a multi-sheet Excel dashboard covering:
  Sheet 1 : Data               — 1,000 synthetic loan records
  Sheet 2 : Dashboard          — 6 analytical sections + Executive KPIs (Actual vs Target)
  Sheet 3 : Portfolio Setup    — Initial assumptions from the Step 0 interview
  Sheet 4 : Pricing Summary    — Key rate table from Loan Pricing.xlsx
  Sheet 5 : Eligibility Rules  — BRE rules from Loan Eligibility Rules.xlsx
  Sheet 6 : Repayment Config   — Assumptions from Repayment Assumptions.xlsx

Run as the FINAL step after all pipeline scripts have completed.

Usage:
    python generate_dashboard.py
    python generate_dashboard.py \\
        --output "C:/path/to/folder" \\
        --pricing "C:/path/Loan Pricing.xlsx" \\
        --rules   "C:/path/Loan Eligibility Rules.xlsx" \\
        --assumptions "C:/path/Repayment Assumptions.xlsx" \\
        --records 3000 --product-weights "40,25,20,15" \\
        --emp-weights "50,30,20" --tier-weights "57,32,11" \\
        --tgt-npa-rate 0.02 --tgt-total-outstanding 79000000
"""

import argparse
import os
import random

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()

# Output
parser.add_argument("--output", default=".", help="Output directory")

# Config files (pipeline outputs to embed)
parser.add_argument("--pricing",     default="", help="Path to Loan Pricing.xlsx")
parser.add_argument("--rules",       default="", help="Path to Loan Eligibility Rules.xlsx")
parser.add_argument("--assumptions", default="", help="Path to Repayment Assumptions.xlsx")

# Initial assumptions (from Step 0 interview — used for Portfolio Setup sheet)
parser.add_argument("--records",           type=int, default=1000)
parser.add_argument("--product-weights",   default="40,25,20,15",
                    help="Personal,Home,Auto,Business weights")
parser.add_argument("--emp-weights",       default="50,30,20",
                    help="Salaried,Self-Employed,Business Owner weights")
parser.add_argument("--tier-weights",      default="",
                    help="Tier1,Tier2,Tier3 %% targets (optional)")
parser.add_argument("--loan-amt-personal", default="50000,1500000")
parser.add_argument("--loan-amt-home",     default="1000000,10000000")
parser.add_argument("--loan-amt-auto",     default="300000,2000000")
parser.add_argument("--loan-amt-business", default="500000,5000000")

# Target KPIs
parser.add_argument("--tgt-total-loans",       type=float, default=0)
parser.add_argument("--tgt-total-disbursed",   type=float, default=0)
parser.add_argument("--tgt-total-outstanding", type=float, default=0)
parser.add_argument("--tgt-npa-count",         type=float, default=0)
parser.add_argument("--tgt-npa-rate",          type=float, default=0)
parser.add_argument("--tgt-delinquent-count",  type=float, default=0)
parser.add_argument("--tgt-portfolio-yield",   type=float, default=0)
parser.add_argument("--tgt-monthly-profit",    type=float, default=0)
parser.add_argument("--tgt-avg-loan-size",     type=float, default=0)
parser.add_argument("--tgt-active-loans",      type=float, default=0)

args = parser.parse_args()
OUT_DIR = args.output
os.makedirs(OUT_DIR, exist_ok=True)


def _parse_weights(s): return [float(x.strip()) for x in s.split(",")]
def _parse_range(s):   p = [int(x.strip()) for x in s.split(",")]; return p[0], p[1]


PRODUCT_WEIGHTS = _parse_weights(args.product_weights)
EMP_WEIGHTS     = _parse_weights(args.emp_weights)
AMT_PERSONAL    = _parse_range(args.loan_amt_personal)
AMT_HOME        = _parse_range(args.loan_amt_home)
AMT_AUTO        = _parse_range(args.loan_amt_auto)
AMT_BUSINESS    = _parse_range(args.loan_amt_business)

# ── SYNTHETIC DATA ────────────────────────────────────────────────────────────
np.random.seed(42)
random.seed(42)

NUM        = args.records
SEGMENTS   = ["Retail", "SME", "Corporate", "Agri"]
PRODUCTS   = ["Personal", "Home", "Auto", "Business"]
AGE_GROUPS = ["18-30", "31-45", "46-60", "60+"]
RATINGS    = ["AAA", "AA", "A", "BBB", "BB", "B", "C"]
RATE_MAP   = {"AAA": 0.07, "AA": 0.08, "A": 0.09, "BBB": 0.11,
              "BB": 0.13, "B": 0.15, "C": 0.18}

rows = []
for i in range(1, NUM + 1):
    seg    = random.choices(SEGMENTS, weights=[50, 30, 10, 10])[0]
    prod   = random.choices(PRODUCTS, weights=PRODUCT_WEIGHTS)[0]
    age    = random.choice(AGE_GROUPS)
    rating = random.choice(RATINGS)

    if prod == "Home":
        disbursed = int(np.random.randint(AMT_HOME[0],     AMT_HOME[1]     + 1))
    elif prod == "Business":
        disbursed = int(np.random.randint(AMT_BUSINESS[0], AMT_BUSINESS[1] + 1))
    elif prod == "Auto":
        disbursed = int(np.random.randint(AMT_AUTO[0],     AMT_AUTO[1]     + 1))
    else:
        disbursed = int(np.random.randint(AMT_PERSONAL[0], AMT_PERSONAL[1] + 1))

    outstanding   = round(disbursed * np.random.uniform(0.1, 0.95), 2)
    interest_rate = round(RATE_MAP[rating] + np.random.uniform(-0.005, 0.005), 4)

    if rating in ["AAA", "AA"]:
        dpd = int(np.random.choice([0, 5, 10], p=[0.95, 0.03, 0.02]))
    elif rating in ["A", "BBB"]:
        dpd = int(np.random.choice([0, 15, 30, 45], p=[0.70, 0.15, 0.10, 0.05]))
    else:
        dpd = int(np.random.choice([0, 45, 90, 120], p=[0.40, 0.30, 0.20, 0.10]))

    status = "NPA" if dpd > 90 else ("Delinquent" if dpd > 0 else "Active")
    profit = round((outstanding * (interest_rate - 0.03)) / 12, 2)

    rows.append([f"L-{1000+i}", seg, prod, age, rating,
                 disbursed, outstanding, interest_rate, status, dpd, profit])

COLS = ["Loan_ID", "Segment", "Product", "Age_Group", "Credit_Rating",
        "Disbursement_Amt", "Outstanding_Principal", "Interest_Rate",
        "Status", "DPD", "Monthly_Profit"]
df = pd.DataFrame(rows, columns=COLS)

# ── PRE-COMPUTE ACTUALS for RAG comparison ────────────────────────────────────
_act_total_loans       = len(df)
_act_total_disbursed   = df["Disbursement_Amt"].sum()
_act_total_outstanding = df["Outstanding_Principal"].sum()
_act_npa_count         = int((df["Status"] == "NPA").sum())
_act_npa_rate          = _act_npa_count / len(df)
_act_delinquent_count  = int((df["Status"] == "Delinquent").sum())
_act_portfolio_yield   = ((df["Interest_Rate"] * df["Outstanding_Principal"]).sum()
                          / df["Outstanding_Principal"].sum())
_act_monthly_profit    = df["Monthly_Profit"].sum()
_act_avg_loan_size     = df["Outstanding_Principal"].sum() / len(df)
_act_active_loans      = int((df["Status"] == "Active").sum())

# ── STYLES ────────────────────────────────────────────────────────────────────
NAVY        = PatternFill("solid", fgColor="1F4E79")
BLUE        = PatternFill("solid", fgColor="2E75B6")
LIGHT_BLUE  = PatternFill("solid", fgColor="D6E4F0")
ALT         = PatternFill("solid", fgColor="EBF5FB")
WHITE       = PatternFill("solid", fgColor="FFFFFF")
AMBER       = PatternFill("solid", fgColor="FFF2CC")
GREEN_RAG   = PatternFill("solid", fgColor="C6EFCE")
RED_RAG     = PatternFill("solid", fgColor="FFC7CE")
GRAY_RAG    = PatternFill("solid", fgColor="F2F2F2")
GREEN_LIGHT = PatternFill("solid", fgColor="E2EFDA")
ORANGE_LIGHT= PatternFill("solid", fgColor="FCE4D6")

THIN_SIDE   = Side(style="thin",   color="AAAAAA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

HDR_FONT   = Font(bold=True, color="FFFFFF",  size=10)
TITLE_FONT = Font(bold=True, color="1F4E79",  size=16)
LABEL_FONT = Font(bold=True, color="1F4E79",  size=10)
KPI_FONT   = Font(bold=True, color="1F4E79",  size=12)
SUB_FONT   = Font(bold=True, color="FFFFFF",  size=10)
NOTE_FONT  = Font(italic=True, color="595959", size=9)
VAL_FONT   = Font(bold=True, color="7B3F00",  size=10)
TOTAL_FONT = Font(bold=True, color="9C0006",  size=10)


def hdr(ws, row, col, value, fill=NAVY, font=HDR_FONT, align=CENTER):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = fill, font, align, THIN_BORDER
    return c


def val(ws, row, col, value, fmt=None, fill=WHITE, align=RIGHT):
    c = ws.cell(row=row, column=col, value=value)
    if fmt:
        c.number_format = fmt
    c.fill, c.alignment, c.border = fill, align, THIN_BORDER
    return c


def section_title(ws, row, title, col_start=1, col_end=6):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=title)
    c.fill      = NAVY
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = THIN_BORDER
    ws.row_dimensions[row].height = 22


def sheet_header(ws, title, subtitle, ncols=6):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font, c.fill, c.alignment, c.border = TITLE_FONT, LIGHT_BLUE, CENTER, THIN_BORDER
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value=subtitle)
    c.font, c.fill, c.border = NOTE_FONT, WHITE, THIN_BORDER
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22


def not_available(ws, msg, R, ncols=6):
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=ncols)
    c = ws.cell(row=R, column=1, value=msg)
    c.font, c.fill, c.alignment, c.border = NOTE_FONT, AMBER, CENTER, THIN_BORDER
    ws.row_dimensions[R].height = 22


# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: DATA
# ═══════════════════════════════════════════════════════════════════════════════
ws_d = wb.active
ws_d.title = "Data"

for ci, name in enumerate(COLS, 1):
    hdr(ws_d, 1, ci, name)

col_fmts_d = {6: "#,##0.00", 7: "#,##0.00", 8: "0.00%", 11: "#,##0.00"}
for ri, row_data in enumerate(df.itertuples(index=False), 2):
    for ci, v in enumerate(row_data, 1):
        c = ws_d.cell(row=ri, column=ci, value=v)
        c.border = THIN_BORDER
        if ci in col_fmts_d:
            c.number_format = col_fmts_d[ci]

for i, w in enumerate([10, 10, 10, 10, 13, 18, 22, 14, 12, 6, 15], 1):
    ws_d.column_dimensions[get_column_letter(i)].width = w
ws_d.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard")
ws.sheet_view.showGridLines = False

for col, w in {1: 28, 2: 20, 3: 20, 4: 16, 5: 18, 6: 16}.items():
    ws.column_dimensions[get_column_letter(col)].width = w

last_row = NUM + 1
SEG  = f"Data!$B$2:$B${last_row}"
PROD = f"Data!$C$2:$C${last_row}"
AGE  = f"Data!$D$2:$D${last_row}"
RAT  = f"Data!$E$2:$E${last_row}"
DIS  = f"Data!$F$2:$F${last_row}"
OUT  = f"Data!$G$2:$G${last_row}"
RATE = f"Data!$H$2:$H${last_row}"
STA  = f"Data!$I$2:$I${last_row}"
DPD_ = f"Data!$J$2:$J${last_row}"
PROF = f"Data!$K$2:$K${last_row}"

# Title
ws.merge_cells("A1:F1")
c = ws["A1"]
c.value, c.font, c.fill = "LOAN PORTFOLIO DASHBOARD", TITLE_FONT, LIGHT_BLUE
c.alignment, c.border = CENTER, THIN_BORDER
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:F2")
c = ws["A2"]
c.value = (f"Generated: {pd.Timestamp.today().strftime('%d %b %Y')}  |  "
           f"Records: {NUM:,}  |  Portfolio simulation with configured assumptions")
c.font, c.fill, c.alignment, c.border = NOTE_FONT, WHITE, CENTER, THIN_BORDER

R = 4

# ── SECTION 1: Executive KPIs — Actual vs Target ──────────────────────────────
section_title(ws, R, "  EXECUTIVE KPIs — ACTUAL vs TARGET")
R += 1

for ci, h in enumerate(["KPI", "Actual", "Target", "vs Target"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 22
R += 1

kpis = [
    ("Total Loans",               f"=COUNTA({SEG})",                          "0",
     args.tgt_total_loans,        _act_total_loans,       True),
    ("Total Disbursed (Rs)",       f"=SUM({DIS})",                             "#,##0.00",
     args.tgt_total_disbursed,    _act_total_disbursed,   True),
    ("Total Outstanding (Rs)",     f"=SUM({OUT})",                             "#,##0.00",
     args.tgt_total_outstanding,  _act_total_outstanding, True),
    ("NPA Count",                  f'=COUNTIF({STA},"NPA")',                   "0",
     args.tgt_npa_count,          _act_npa_count,         False),
    ("NPA Rate (%)",               f'=COUNTIF({STA},"NPA")/COUNTA({SEG})',     "0.00%",
     args.tgt_npa_rate,           _act_npa_rate,          False),
    ("Delinquent Count",           f'=COUNTIF({STA},"Delinquent")',            "0",
     args.tgt_delinquent_count,   _act_delinquent_count,  False),
    ("Portfolio Yield (%)",        f"=SUMPRODUCT({RATE},{OUT})/SUM({OUT})",    "0.00%",
     args.tgt_portfolio_yield,    _act_portfolio_yield,   True),
    ("Total Monthly Profit (Rs)",  f"=SUM({PROF})",                            "#,##0.00",
     args.tgt_monthly_profit,     _act_monthly_profit,    True),
    ("Avg Loan Size (Rs)",         f"=SUM({OUT})/COUNTA({SEG})",               "#,##0.00",
     args.tgt_avg_loan_size,      _act_avg_loan_size,     True),
    ("Active Loans",               f'=COUNTIF({STA},"Active")',                "0",
     args.tgt_active_loans,       _act_active_loans,      True),
]

for label, formula, fmt, target, actual, higher_is_better in kpis:
    lc = ws.cell(row=R, column=1, value=label)
    lc.font, lc.fill, lc.alignment, lc.border = LABEL_FONT, LIGHT_BLUE, LEFT, THIN_BORDER

    ac = ws.cell(row=R, column=2, value=formula)
    ac.number_format = fmt
    ac.font, ac.fill, ac.alignment, ac.border = KPI_FONT, WHITE, CENTER, THIN_BORDER

    tc = ws.cell(row=R, column=3, value=target if target > 0 else None)
    tc.number_format = fmt
    tc.fill, tc.alignment, tc.border = AMBER, CENTER, THIN_BORDER
    tc.font = Font(bold=True, color="7B3F00", size=11)

    if target > 0:
        on_track = (actual >= target) if higher_is_better else (actual <= target)
        sc = ws.cell(row=R, column=4, value="▲ On Track" if on_track else "▼ Off Target")
        sc.fill = GREEN_RAG if on_track else RED_RAG
        sc.font = Font(bold=True, color="375623" if on_track else "9C0006", size=10)
    else:
        sc = ws.cell(row=R, column=4, value="— No Target Set")
        sc.fill = GRAY_RAG
        sc.font = Font(italic=True, color="595959", size=9)
    sc.alignment, sc.border = CENTER, THIN_BORDER
    ws.row_dimensions[R].height = 20
    R += 1

R += 1

# ── SECTION 2: NPA by Segment ─────────────────────────────────────────────────
section_title(ws, R, "  NPA BY SEGMENT")
R += 1
for ci, h in enumerate(["Segment", "Total Loans", "NPA Count",
                         "NPA %", "Outstanding (NPA Rs)", "Delinquent Count"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1
for i, seg in enumerate(SEGMENTS):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, seg,                                                      fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({SEG},"{seg}")',             "#,##0",          fill=fill)
    val(ws, R, 3, f'=COUNTIFS({SEG},"{seg}",{STA},"NPA")',"#,##0",         fill=fill)
    val(ws, R, 4, f'=IFERROR(COUNTIFS({SEG},"{seg}",{STA},"NPA")'
                  f'/COUNTIF({SEG},"{seg}"),0)',           "0.00%",          fill=fill)
    val(ws, R, 5, f'=SUMIFS({OUT},{SEG},"{seg}",{STA},"NPA")', "#,##0.00", fill=fill)
    val(ws, R, 6, f'=COUNTIFS({SEG},"{seg}",{STA},"Delinquent")', "#,##0", fill=fill)
    R += 1
R += 1

# ── SECTION 3: Portfolio Concentration by Product ─────────────────────────────
section_title(ws, R, "  PORTFOLIO CONCENTRATION BY PRODUCT")
R += 1
for ci, h in enumerate(["Product", "Count", "Outstanding (Rs)",
                         "% of Portfolio", "NPA Count", "Avg Interest Rate"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1
for i, prod in enumerate(PRODUCTS):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, prod,                                                         fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({PROD},"{prod}")',              "#,##0",           fill=fill)
    val(ws, R, 3, f'=SUMIF({PROD},"{prod}",{OUT})',          "#,##0.00",        fill=fill)
    val(ws, R, 4, f'=IFERROR(SUMIF({PROD},"{prod}",{OUT})/SUM({OUT}),0)', "0.00%", fill=fill)
    val(ws, R, 5, f'=COUNTIFS({PROD},"{prod}",{STA},"NPA")', "#,##0",          fill=fill)
    val(ws, R, 6, f'=AVERAGEIF({PROD},"{prod}",{RATE})',     "0.00%",           fill=fill)
    R += 1
R += 1

# ── SECTION 4: Age Group vs DPD ───────────────────────────────────────────────
section_title(ws, R, "  AGE GROUP vs DPD ANALYSIS")
R += 1
for ci, h in enumerate(["Age Group", "Count", "Avg DPD",
                         "Delinquent Count", "NPA Count", "NPA %"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1
for i, age in enumerate(AGE_GROUPS):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, age,                                                            fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({AGE},"{age}")',                "#,##0",            fill=fill)
    val(ws, R, 3, f'=AVERAGEIF({AGE},"{age}",{DPD_})',       "0.0",              fill=fill)
    val(ws, R, 4, f'=COUNTIFS({AGE},"{age}",{STA},"Delinquent")', "#,##0",      fill=fill)
    val(ws, R, 5, f'=COUNTIFS({AGE},"{age}",{STA},"NPA")',   "#,##0",            fill=fill)
    val(ws, R, 6, f'=IFERROR(COUNTIFS({AGE},"{age}",{STA},"NPA")'
                  f'/COUNTIF({AGE},"{age}"),0)',              "0.00%",            fill=fill)
    R += 1
R += 1

# ── SECTION 5: Interest Rate Band Analysis ────────────────────────────────────
section_title(ws, R, "  INTEREST RATE BAND ANALYSIS")
R += 1
for ci, h in enumerate(["Rate Band", "Count", "Avg Rate",
                         "Total Outstanding (Rs)", "Total Monthly Profit (Rs)", "NPA Count"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1
bands = [
    ("Low  (< 8%)",
     f"=SUMPRODUCT(({RATE}<0.08)*1)",
     f"=IFERROR(SUMPRODUCT(({RATE}<0.08)*{RATE})/SUMPRODUCT(({RATE}<0.08)*1),0)",
     f"=SUMPRODUCT(({RATE}<0.08)*{OUT})",
     f"=SUMPRODUCT(({RATE}<0.08)*{PROF})",
     f'=SUMPRODUCT(({RATE}<0.08)*({STA}="NPA"))'),
    ("Medium (8-12%)",
     f"=SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*1)",
     f"=IFERROR(SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*{RATE})"
     f"/SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*1),0)",
     f"=SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*{OUT})",
     f"=SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*{PROF})",
     f'=SUMPRODUCT(({RATE}>=0.08)*({RATE}<0.12)*({STA}="NPA"))'),
    ("High  (>=12%)",
     f"=SUMPRODUCT(({RATE}>=0.12)*1)",
     f"=IFERROR(SUMPRODUCT(({RATE}>=0.12)*{RATE})/SUMPRODUCT(({RATE}>=0.12)*1),0)",
     f"=SUMPRODUCT(({RATE}>=0.12)*{OUT})",
     f"=SUMPRODUCT(({RATE}>=0.12)*{PROF})",
     f'=SUMPRODUCT(({RATE}>=0.12)*({STA}="NPA"))'),
]
for i, (label, cnt, avg_r, tot_out, tot_prof, npa_cnt) in enumerate(bands):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, label,    fill=fill, align=LEFT)
    val(ws, R, 2, cnt,      "#,##0",    fill=fill)
    val(ws, R, 3, avg_r,    "0.00%",    fill=fill)
    val(ws, R, 4, tot_out,  "#,##0.00", fill=fill)
    val(ws, R, 5, tot_prof, "#,##0.00", fill=fill)
    val(ws, R, 6, npa_cnt,  "#,##0",    fill=fill)
    R += 1
R += 1

# ── SECTION 6: Profitability by Credit Rating ─────────────────────────────────
section_title(ws, R, "  PROFITABILITY BY CREDIT RATING")
R += 1
for ci, h in enumerate(["Credit Rating", "Count", "Avg Interest Rate",
                         "Avg Monthly Profit (Rs)", "Total Monthly Profit (Rs)", "NPA Count"], 1):
    hdr(ws, R, ci, h, fill=BLUE, font=SUB_FONT)
ws.row_dimensions[R].height = 28
R += 1
for i, rating in enumerate(RATINGS):
    fill = ALT if i % 2 == 0 else WHITE
    val(ws, R, 1, rating,                                                           fill=fill, align=LEFT)
    val(ws, R, 2, f'=COUNTIF({RAT},"{rating}")',              "#,##0",              fill=fill)
    val(ws, R, 3, f'=AVERAGEIF({RAT},"{rating}",{RATE})',     "0.00%",              fill=fill)
    val(ws, R, 4, f'=AVERAGEIF({RAT},"{rating}",{PROF})',     "#,##0.00",           fill=fill)
    val(ws, R, 5, f'=SUMIF({RAT},"{rating}",{PROF})',         "#,##0.00",           fill=fill)
    val(ws, R, 6, f'=COUNTIFS({RAT},"{rating}",{STA},"NPA")', "#,##0",             fill=fill)
    R += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: PORTFOLIO SETUP — INITIAL ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Portfolio Setup")
ws_s.sheet_view.showGridLines = False

for col, w in {1: 30, 2: 22, 3: 22, 4: 22, 5: 22, 6: 20}.items():
    ws_s.column_dimensions[get_column_letter(col)].width = w

sheet_header(ws_s, "PORTFOLIO SETUP — INITIAL ASSUMPTIONS",
             "Configuration used in the Step 0 interview to generate the loan applications data.")

R = 4

# Section: Records
section_title(ws_s, R, "  APPLICATION VOLUME", col_end=4)
R += 1
hdr(ws_s, R, 1, "Parameter", fill=BLUE, font=SUB_FONT, align=LEFT)
hdr(ws_s, R, 2, "Value",     fill=BLUE, font=SUB_FONT)
ws_s.row_dimensions[R].height = 22
R += 1
val(ws_s, R, 1, "Total Records Generated", fill=LIGHT_BLUE, align=LEFT)
c = ws_s.cell(row=R, column=2, value=NUM)
c.fill, c.font, c.alignment, c.border = AMBER, VAL_FONT, CENTER, THIN_BORDER
c.number_format = "#,##0"
R += 2

# Section: Product Distribution
section_title(ws_s, R, "  LOAN PRODUCT DISTRIBUTION", col_end=4)
R += 1
for ci, h in enumerate(["Product", "Configured Weight", "Actual Count", "Actual %"], 1):
    hdr(ws_s, R, ci, h, fill=BLUE, font=SUB_FONT)
ws_s.row_dimensions[R].height = 22
R += 1
total_w = sum(PRODUCT_WEIGHTS)
for i, (prod, wt) in enumerate(zip(PRODUCTS, PRODUCT_WEIGHTS)):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    act_cnt = int((df["Product"] == prod).sum())
    val(ws_s, R, 1, prod,               fill=fill, align=LEFT)
    c = ws_s.cell(row=R, column=2, value=round(wt / total_w * 100, 1))
    c.fill, c.font, c.number_format, c.alignment, c.border = AMBER, VAL_FONT, "0.0", CENTER, THIN_BORDER
    val(ws_s, R, 3, act_cnt,            "#,##0",   fill=fill)
    val(ws_s, R, 4, act_cnt / NUM,      "0.00%",   fill=fill)
    R += 1
R += 1

# Section: Employment Distribution
section_title(ws_s, R, "  CUSTOMER EMPLOYMENT TYPE", col_end=4)
R += 1
for ci, h in enumerate(["Employment Type", "Configured Weight", "Description", ""], 1):
    hdr(ws_s, R, ci, h, fill=BLUE, font=SUB_FONT)
ws_s.row_dimensions[R].height = 22
R += 1
emp_labels = ["Salaried", "Self-Employed", "Business Owner"]
emp_descs  = ["Regular salary from employer", "Independent professional or freelancer",
              "Owns and operates a business"]
total_ew = sum(EMP_WEIGHTS)
for i, (emp, wt, desc) in enumerate(zip(emp_labels, EMP_WEIGHTS, emp_descs)):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    val(ws_s, R, 1, emp,               fill=fill, align=LEFT)
    c = ws_s.cell(row=R, column=2, value=round(wt / total_ew * 100, 1))
    c.fill, c.font, c.number_format, c.alignment, c.border = AMBER, VAL_FONT, "0.0", CENTER, THIN_BORDER
    val(ws_s, R, 3, desc,              fill=fill, align=LEFT)
    val(ws_s, R, 4, "",                fill=fill)
    R += 1
R += 1

# Section: City Tier Distribution
section_title(ws_s, R, "  CITY TIER DISTRIBUTION", col_end=4)
R += 1
for ci, h in enumerate(["City Tier", "Configured Target %%", "Cities Included", ""], 1):
    hdr(ws_s, R, ci, h, fill=BLUE, font=SUB_FONT)
ws_s.row_dimensions[R].height = 22
R += 1
if args.tier_weights:
    tw = _parse_weights(args.tier_weights)
    tier_rows = [("Tier 1", tw[0], "8 metros: Mumbai, Delhi, Bengaluru, Chennai, Hyderabad, Kolkata, Pune, Ahmedabad"),
                 ("Tier 2", tw[1], "14 cities: Jaipur, Lucknow, Surat, Kochi, Vadodara, etc."),
                 ("Tier 3", tw[2], "13 cities: Madurai, Varanasi, Agra, Mysuru, Guwahati, etc.")]
else:
    tier_rows = [("Tier 1", "Default (~57%)", "8 metros: Mumbai, Delhi, Bengaluru, Chennai, Hyderabad, Kolkata, Pune, Ahmedabad"),
                 ("Tier 2", "Default (~32%)", "14 cities: Jaipur, Lucknow, Surat, Kochi, Vadodara, etc."),
                 ("Tier 3", "Default (~11%)", "13 cities: Madurai, Varanasi, Agra, Mysuru, Guwahati, etc.")]
for i, (tier, tgt, cities) in enumerate(tier_rows):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    val(ws_s, R, 1, tier,   fill=fill, align=LEFT)
    c = ws_s.cell(row=R, column=2, value=tgt)
    c.fill, c.font, c.alignment, c.border = AMBER, VAL_FONT, CENTER, THIN_BORDER
    val(ws_s, R, 3, cities, fill=fill, align=LEFT)
    val(ws_s, R, 4, "",     fill=fill)
    R += 1
R += 1

# Section: Loan Amount Ranges
section_title(ws_s, R, "  LOAN AMOUNT RANGES (Rs)", col_end=4)
R += 1
for ci, h in enumerate(["Product", "Minimum (Rs)", "Maximum (Rs)", "Range"], 1):
    hdr(ws_s, R, ci, h, fill=BLUE, font=SUB_FONT)
ws_s.row_dimensions[R].height = 22
R += 1
amt_rows = [("Personal", AMT_PERSONAL), ("Home", AMT_HOME),
            ("Auto", AMT_AUTO),         ("Business", AMT_BUSINESS)]
for i, (prod, (lo, hi)) in enumerate(amt_rows):
    fill = GREEN_LIGHT if i % 2 == 0 else WHITE
    val(ws_s, R, 1, prod,                      fill=fill, align=LEFT)
    c2 = ws_s.cell(row=R, column=2, value=lo)
    c2.fill, c2.font, c2.number_format = AMBER, VAL_FONT, "#,##0"
    c2.alignment, c2.border = CENTER, THIN_BORDER
    c3 = ws_s.cell(row=R, column=3, value=hi)
    c3.fill, c3.font, c3.number_format = AMBER, VAL_FONT, "#,##0"
    c3.alignment, c3.border = CENTER, THIN_BORDER
    val(ws_s, R, 4, f"Rs {lo:,.0f}  to  Rs {hi:,.0f}", fill=fill, align=LEFT)
    R += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: PRICING SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
ws_p = wb.create_sheet("Pricing Summary")
ws_p.sheet_view.showGridLines = False

for col, w in {1: 22, 2: 18, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16}.items():
    ws_p.column_dimensions[get_column_letter(col)].width = w

sheet_header(ws_p, "LOAN PRICING SUMMARY",
             "Interest rate spread logic and average rates by risk dimension.",
             ncols=7)

R = 4

pricing_file = args.pricing or os.path.join(OUT_DIR, "Loan Pricing.xlsx")
if os.path.exists(pricing_file):
    try:
        pricing_raw = pd.read_excel(pricing_file, sheet_name="Pricing Table", header=8)
        pricing_raw.columns = [str(c).strip() for c in pricing_raw.columns]
        # Drop blank/unnamed columns (col A is empty in the pricing sheet)
        pricing_raw = pricing_raw[[c for c in pricing_raw.columns
                                    if not c.startswith("Unnamed")]]

        # ── Spread Logic summary ──────────────────────────────────────────────
        section_title(ws_p, R, "  SPREAD LOGIC — BASE RATE + ADDITIVE SPREADS", col_end=7)
        R += 1
        for ci, h in enumerate(["Component", "Min Spread %%", "Max Spread %%",
                                 "Driver", "", "", ""], 1):
            hdr(ws_p, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_p.row_dimensions[R].height = 22
        R += 1
        spread_rows = [
            ("Base Rate",          "9.00%", "9.00%", "MCLR / Repo-linked rate — adjust to policy"),
            ("CIBIL Spread",       "0.50%", "4.00%", "Higher risk (lower CIBIL) → higher spread"),
            ("Employment Spread",  "0.00%", "1.00%", "Govt/PSU = 0, Business Owner = max"),
            ("Loan Amount Spread", "0.00%", "0.50%", "Larger loans cheaper — volume incentive"),
            ("Tenure Spread",      "-0.25%","0.25%", "Longer tenure = slightly higher rate"),
            ("Age Spread",         "0.00%", "0.25%", "Near-retirement (46-50) small premium"),
        ]
        for i, (comp, lo, hi, note) in enumerate(spread_rows):
            fill = GREEN_LIGHT if i % 2 == 0 else WHITE
            val(ws_p, R, 1, comp, fill=fill, align=LEFT)
            val(ws_p, R, 2, lo,   fill=AMBER if i > 0 else fill, align=CENTER)
            val(ws_p, R, 3, hi,   fill=AMBER if i > 0 else fill, align=CENTER)
            ws_p.merge_cells(start_row=R, start_column=4, end_row=R, end_column=7)
            val(ws_p, R, 4, note, fill=fill, align=LEFT)
            R += 1
        R += 1

        # ── Avg rate pivot: CIBIL band × Employment ───────────────────────────
        section_title(ws_p, R, "  AVG FINAL RATE (%) BY CIBIL BAND x EMPLOYMENT TYPE", col_end=7)
        R += 1

        # Find the rate column
        rate_col  = next((c for c in pricing_raw.columns
                          if "final" in c.lower() and "rate" in c.lower()), None)
        emp_col   = next((c for c in pricing_raw.columns
                          if "employment" in c.lower() and "type" in c.lower()), None)
        cibil_col = next((c for c in pricing_raw.columns
                          if "cibil" in c.lower()), None)   # "CIBIL Score Range"

        if rate_col and emp_col and cibil_col:
            pivot = (pricing_raw.groupby([cibil_col, emp_col])[rate_col]
                     .mean().unstack(fill_value=0))
            emp_cols = list(pivot.columns)

            # Header row
            hdr(ws_p, R, 1, "CIBIL Band", fill=BLUE, font=SUB_FONT, align=LEFT)
            for ci, ec in enumerate(emp_cols, 2):
                hdr(ws_p, R, ci, str(ec), fill=BLUE, font=SUB_FONT)
            ws_p.row_dimensions[R].height = 28
            R += 1

            for i, (cibil, row_data) in enumerate(pivot.iterrows()):
                fill = ALT if i % 2 == 0 else WHITE
                val(ws_p, R, 1, str(cibil), fill=fill, align=LEFT)
                for ci, rate in enumerate(row_data, 2):
                    c = ws_p.cell(row=R, column=ci, value=round(rate, 2) if rate else None)
                    c.number_format = "0.00"
                    c.fill, c.alignment, c.border = fill, CENTER, THIN_BORDER
                R += 1
        else:
            not_available(ws_p, "Could not parse Pricing Table columns — check sheet format.", R, 7)
            R += 1

    except Exception as e:
        not_available(ws_p, f"Could not read Loan Pricing.xlsx: {e}", R, 7)
        R += 1
else:
    not_available(ws_p,
                  "Loan Pricing.xlsx not found. Run build_loan_pricing.py and pass --pricing path.",
                  R, 7)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: ELIGIBILITY RULES
# ═══════════════════════════════════════════════════════════════════════════════
ws_e = wb.create_sheet("Eligibility Rules")
ws_e.sheet_view.showGridLines = False

for col, w in {1: 10, 2: 28, 3: 24, 4: 14, 5: 28, 6: 22, 7: 10, 8: 38}.items():
    ws_e.column_dimensions[get_column_letter(col)].width = w

sheet_header(ws_e, "LOAN ELIGIBILITY RULES (BRE)",
             "Business Rules Engine configuration used to filter approved loans.",
             ncols=8)

R = 4
rules_file = args.rules or os.path.join(OUT_DIR, "Loan Eligibility Rules.xlsx")
if os.path.exists(rules_file):
    try:
        rules_df = pd.read_excel(rules_file, sheet_name="Eligibility Rules", header=3)
        rules_df.columns = [str(c).strip().replace(" \u2710", "") for c in rules_df.columns]

        # Column headers
        display_cols = ["Rule ID", "Rule Name", "Field", "Operator",
                        "Value", "Applies To", "Enabled", "Description"]
        avail = [c for c in display_cols if c in rules_df.columns]
        for ci, h in enumerate(avail, 1):
            hdr(ws_e, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_e.row_dimensions[R].height = 28
        R += 1

        for i, row in rules_df.iterrows():
            if pd.isna(row.get("Rule ID", None)):
                continue
            fill = GREEN_LIGHT if i % 2 == 0 else WHITE
            enabled = str(row.get("Enabled", "Yes")).strip().lower()
            if enabled != "yes":
                fill = GRAY_RAG
            for ci, col_name in enumerate(avail, 1):
                v = row.get(col_name, "")
                c = ws_e.cell(row=R, column=ci,
                              value=str(v) if not pd.isna(v) else "")
                # Amber highlight on Value column
                c.fill = AMBER if col_name == "Value" else fill
                c.font = VAL_FONT if col_name == "Value" else Font(size=10)
                c.alignment = LEFT if col_name in ("Rule Name", "Description", "Applies To") else CENTER
                c.border = THIN_BORDER
            R += 1

        # Legend
        R += 1
        ws_e.merge_cells(start_row=R, start_column=1, end_row=R, end_column=8)
        c = ws_e.cell(row=R, column=1,
                      value="Amber = editable Value column  |  Grey rows = rule disabled  |  "
                            "Edit in Loan Eligibility Rules.xlsx then re-run generate_approved_loans.py")
        c.font, c.fill, c.alignment, c.border = NOTE_FONT, AMBER, LEFT, THIN_BORDER

    except Exception as e:
        not_available(ws_e, f"Could not read Loan Eligibility Rules.xlsx: {e}", R, 8)
else:
    not_available(ws_e,
                  "Loan Eligibility Rules.xlsx not found. Run build_loan_eligibility_rules.py.",
                  R, 8)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: REPAYMENT CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
ws_r = wb.create_sheet("Repayment Config")
ws_r.sheet_view.showGridLines = False

for col, w in {1: 38, 2: 16, 3: 14, 4: 36}.items():
    ws_r.column_dimensions[get_column_letter(col)].width = w

sheet_header(ws_r, "REPAYMENT SCHEDULE — ASSUMPTIONS",
             "Configuration used to simulate repayment behaviour, DPD, prepayments and penalties.",
             ncols=4)

R = 4

assump_file = args.assumptions or os.path.join(OUT_DIR, "Repayment Assumptions.xlsx")
if os.path.exists(assump_file):
    try:
        wb_a = load_workbook(assump_file, data_only=True)
        ws_a = wb_a["Repayment Assumptions"]

        def _cell(r, c):
            v = ws_a.cell(row=r, column=c).value
            return v if v is not None else ""

        # ── Section 1: Payment Profile Distribution ───────────────────────────
        section_title(ws_r, R, "  SECTION 1 — PAYMENT PROFILE DISTRIBUTION", col_end=4)
        R += 1
        for ci, h in enumerate(["Profile", "% of Loans", "Description", ""], 1):
            hdr(ws_r, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_r.row_dimensions[R].height = 22
        R += 1
        # Rows 6-10 in the assumptions sheet
        profile_rows = [
            (6,  "Always pays on time. DPD = 0 every month."),
            (7,  "Mostly on time, misses occasionally (see DPD Occasional section)."),
            (8,  "Frequently late. DPD pattern per Chronic section."),
            (9,  "On time, makes one partial prepayment mid-tenure."),
            (10, "On time, closes loan early in full."),
        ]
        total_pct = 0
        for idx, (src_row, desc) in enumerate(profile_rows):
            name = _cell(src_row, 1)
            pct  = _cell(src_row, 2)
            fill = GREEN_LIGHT if idx % 2 == 0 else WHITE
            val(ws_r, R, 1, name, fill=fill, align=LEFT)
            c = ws_r.cell(row=R, column=2, value=pct)
            c.fill, c.font, c.number_format = AMBER, VAL_FONT, "0.00"
            c.alignment, c.border = CENTER, THIN_BORDER
            val(ws_r, R, 3, desc, fill=fill, align=LEFT)
            val(ws_r, R, 4, "",   fill=fill)
            if isinstance(pct, (int, float)):
                total_pct += pct
            R += 1
        # Total
        val(ws_r, R, 1, "TOTAL", fill=ORANGE_LIGHT, align=LEFT)
        c = ws_r.cell(row=R, column=2, value=total_pct)
        c.fill, c.font, c.number_format = ORANGE_LIGHT, TOTAL_FONT, "0.00"
        c.alignment, c.border = CENTER, THIN_BORDER
        val(ws_r, R, 3, "Must equal 100", fill=ORANGE_LIGHT, align=LEFT)
        val(ws_r, R, 4, "", fill=ORANGE_LIGHT)
        R += 2

        # ── Section 2: DPD — Occasional Delay ────────────────────────────────
        section_title(ws_r, R, "  SECTION 2 — DPD: OCCASIONAL DELAY", col_end=4)
        R += 1
        for ci, h in enumerate(["DPD Scenario", "DPD Days", "Probability %%", "Description"], 1):
            hdr(ws_r, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_r.row_dimensions[R].height = 22
        R += 1
        for idx, src_row in enumerate([15, 16, 17]):
            fill = GREEN_LIGHT if idx % 2 == 0 else WHITE
            val(ws_r, R, 1, _cell(src_row, 1), fill=fill, align=LEFT)
            for ci in (2, 3):
                c = ws_r.cell(row=R, column=ci, value=_cell(src_row, ci))
                c.fill, c.font, c.number_format = AMBER, VAL_FONT, "0.00"
                c.alignment, c.border = CENTER, THIN_BORDER
            val(ws_r, R, 4, _cell(src_row, 4), fill=fill, align=LEFT)
            R += 1
        R += 1

        # ── Section 3: DPD — Chronic Delay ───────────────────────────────────
        section_title(ws_r, R, "  SECTION 3 — DPD: CHRONIC DELAY", col_end=4)
        R += 1
        for ci, h in enumerate(["DPD Scenario", "DPD Days", "Probability %%", "Description"], 1):
            hdr(ws_r, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_r.row_dimensions[R].height = 22
        R += 1
        for idx, src_row in enumerate([22, 23, 24, 25, 26]):
            fill = GREEN_LIGHT if idx % 2 == 0 else WHITE
            val(ws_r, R, 1, _cell(src_row, 1), fill=fill, align=LEFT)
            for ci in (2, 3):
                c = ws_r.cell(row=R, column=ci, value=_cell(src_row, ci))
                c.fill, c.font, c.number_format = AMBER, VAL_FONT, "0.00"
                c.alignment, c.border = CENTER, THIN_BORDER
            val(ws_r, R, 4, _cell(src_row, 4), fill=fill, align=LEFT)
            R += 1
        R += 1

        # ── Sections 4-6: Key parameters ─────────────────────────────────────
        section_title(ws_r, R, "  SECTIONS 4-6 — PREPAYMENT & PENAL PARAMETERS", col_end=4)
        R += 1
        for ci, h in enumerate(["Parameter", "Value", "Unit", "Description"], 1):
            hdr(ws_r, R, ci, h, fill=BLUE, font=SUB_FONT)
        ws_r.row_dimensions[R].height = 22
        R += 1

        # Section 4: rows 31-34; Section 5: rows 38-40; Section 6: rows 44-48
        param_sections = [
            ("Partial Prepayment",  [31, 32, 33, 34]),
            ("Full Prepayment",     [38, 39, 40]),
            ("Write-offs & Penal",  [44, 45, 46, 47, 48]),
        ]
        row_idx = 0
        for sec_name, src_rows in param_sections:
            # Sub-label
            ws_r.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
            c = ws_r.cell(row=R, column=1, value=f"  {sec_name}")
            c.fill = LIGHT_BLUE
            c.font = Font(bold=True, color="1F4E79", size=10)
            c.alignment, c.border = LEFT, THIN_BORDER
            R += 1
            for src_row in src_rows:
                fill = GREEN_LIGHT if row_idx % 2 == 0 else WHITE
                val(ws_r, R, 1, _cell(src_row, 1), fill=fill, align=LEFT)
                c = ws_r.cell(row=R, column=2, value=_cell(src_row, 2))
                c.fill, c.font, c.number_format = AMBER, VAL_FONT, "0.00"
                c.alignment, c.border = CENTER, THIN_BORDER
                val(ws_r, R, 3, _cell(src_row, 3), fill=fill, align=CENTER)
                val(ws_r, R, 4, _cell(src_row, 4), fill=fill, align=LEFT)
                R += 1
                row_idx += 1

    except Exception as e:
        not_available(ws_r, f"Could not read Repayment Assumptions.xlsx: {e}", R, 4)
else:
    not_available(ws_r,
                  "Repayment Assumptions.xlsx not found. Run build_repayment_assumptions.py.",
                  R, 4)


# ── SAVE ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(OUT_DIR, "loan_portfolio_dashboard.xlsx")
wb.save(out_path)
print(f"Dashboard saved: {out_path}")
print(f"  Sheets: Data | Dashboard | Portfolio Setup | Pricing Summary | "
      f"Eligibility Rules | Repayment Config")
