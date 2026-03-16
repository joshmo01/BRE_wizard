"""
Repayment Schedule Dashboard Generator
Reads loan_repayment_schedule.xlsx and builds a formula-driven Excel dashboard.

Columns in Repayment_Schedule sheet (A-U):
  A  Schedule_ID      B  Loan_ID          C  EMI_Number
  D  Due_Date         E  Payment_Date     F  DPD
  G  DPD_Bucket       H  Opening_Balance  I  Scheduled_EMI
  J  Principal_Comp   K  Interest_Comp    L  Penal_Interest
  M  Prepayment_Amt   N  Prepayment_Chg   O  Prepayment_Type
  P  Default_Penalty  Q  Loan_Status      R  Closing_Balance
  S  Payment_Status   T  Total_Amt_Due    U  Total_Amt_Paid

Usage:
  python generate_repayment_dashboard.py
  python generate_repayment_dashboard.py --input "path/loan_repayment_schedule.xlsx"
                                          --output "path/to/folder"
"""

import argparse
import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--input",  default=r"C:\Users\joshm\OneDrive\Documents\BRE\loan_repayment_schedule.xlsx")
parser.add_argument("--output", default=r"C:\Users\joshm\OneDrive\Documents\BRE")
args = parser.parse_args()

OUT_PATH = os.path.join(args.output, "repayment_dashboard.xlsx")
os.makedirs(args.output, exist_ok=True)

# ── LOAD & INSPECT ────────────────────────────────────────────────────────────
df = pd.read_excel(args.input, sheet_name="Repayment_Schedule")
LAST_ROW = len(df) + 1          # header is row 1; data ends at len(df)+1
print(f"Loaded {len(df):,} schedule rows  (data rows 2-{LAST_ROW})")

# ── OPEN WORKBOOK AND ADD DASHBOARD SHEET ─────────────────────────────────────
wb = load_workbook(args.input)
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws = wb.create_sheet("Dashboard")
wb.move_sheet("Dashboard", offset=-len(wb.sheetnames) + 1)

DATA = "Repayment_Schedule"
LR   = LAST_ROW

def rng(col):
    """Returns a bare range expression (no leading =) for embedding in formulas."""
    return f"'{DATA}'!${col}$2:${col}${LR}"

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
ALT_ROW     = "EBF3FB"
WHITE       = "FFFFFF"

def hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

def section_header(ws, row, col, text, span=6, bg=DARK_BLUE, fg="FFFFFF", size=11):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Calibri", bold=True, size=size, color=fg)
    cell.fill      = hdr_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)

def col_header(ws, row, col, text, bg=MID_BLUE, fg="FFFFFF"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Calibri", bold=True, size=9, color=fg)
    cell.fill      = hdr_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin_border()

def kpi_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
    c.fill      = hdr_fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = thin_border()

def kpi_value(ws, row, col, expr, fmt="General"):
    """expr must NOT include a leading = — it is prepended here."""
    c = ws.cell(row=row, column=col, value="=" + expr)
    c.font        = Font(name="Calibri", bold=True, size=10, color=DARK_BLUE)
    c.fill        = hdr_fill(WHITE)
    c.alignment   = Alignment(horizontal="right", vertical="center")
    c.border      = thin_border()
    c.number_format = fmt

def data_cell(ws, row, col, expr, fmt="General", alt=False):
    """expr must NOT include a leading = — it is prepended here."""
    c = ws.cell(row=row, column=col, value="=" + expr)
    c.font        = Font(name="Calibri", size=9)
    c.fill        = hdr_fill(ALT_ROW if alt else WHITE)
    c.alignment   = Alignment(horizontal="right", vertical="center")
    c.border      = thin_border()
    c.number_format = fmt

def row_label(ws, row, col, text, alt=False):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Calibri", bold=True, size=9, color=DARK_BLUE)
    c.fill      = hdr_fill(ALT_ROW if alt else WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = thin_border()

# ── COLUMN WIDTHS ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 2
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 18
ws.column_dimensions["G"].width = 18
ws.column_dimensions["H"].width = 2

ws.row_dimensions[1].height = 6

# ── TITLE ─────────────────────────────────────────────────────────────────────
section_header(ws, 2, 2, "LOAN REPAYMENT SCHEDULE DASHBOARD", span=6,
               bg=DARK_BLUE, size=14)
ws.row_dimensions[2].height = 30

# =============================================================================
# SECTION 1 — EXECUTIVE KPIs  (rows 4-10)
# =============================================================================
section_header(ws, 4, 2, "1. EXECUTIVE KPIs", span=6, bg=MID_BLUE)
ws.row_dimensions[4].height = 22

kpis_left = [
    ("Total EMI Records",       f"COUNTA({rng('A')})",                                     "#,##0"),
    ("Total Loans",             f"SUMPRODUCT(1/COUNTIF({rng('B')},{rng('B')}))",            "#,##0"),
    ("Total Scheduled (Rs)",    f"SUM({rng('I')})",                                         "#,##0"),
    ("Total Collected (Rs)",    f"SUM({rng('U')})",                                         "#,##0"),
    ("Collection Efficiency %", f"SUM({rng('U')})/SUM({rng('T')})*100",                     "0.00\"%\""),
]
kpis_right = [
    ("Total Penal Interest (Rs)",   f"SUM({rng('L')})",                         "#,##0"),
    ("Total Default Penalty (Rs)",  f"SUM({rng('P')})",                         "#,##0"),
    ("Total Prepayment Chg (Rs)",   f"SUM({rng('N')})",                         "#,##0"),
    ("NPA EMI Rows",                f"COUNTIF({rng('Q')},\"NPA\")",              "#,##0"),
    ("Avg DPD (all EMIs)",          f"AVERAGE({rng('F')})",                      "0.00"),
]

for i, (label, expr, fmt) in enumerate(kpis_left):
    r = 5 + i
    ws.row_dimensions[r].height = 20
    kpi_label(ws, r, 2, label)
    kpi_value(ws, r, 3, expr, fmt)

for i, (label, expr, fmt) in enumerate(kpis_right):
    r = 5 + i
    kpi_label(ws, r, 5, label)
    kpi_value(ws, r, 6, expr, fmt)

# =============================================================================
# SECTION 2 — PAYMENT STATUS DISTRIBUTION  (rows 12-19)
# =============================================================================
SEC2 = 12
section_header(ws, SEC2, 2, "2. PAYMENT STATUS DISTRIBUTION", span=6, bg=MID_BLUE)
ws.row_dimensions[SEC2].height = 22

for c, h in enumerate(["Payment Status", "EMI Count", "% of Total",
                        "Total Paid (Rs)", "Avg DPD", "Penal Interest (Rs)"], 2):
    col_header(ws, SEC2 + 1, c, h)
ws.row_dimensions[SEC2 + 1].height = 30

total_emi_expr = f"COUNTA({rng('S')})"
for i, status in enumerate(["On Time", "Delayed", "Defaulted", "Part Prepay", "Full Prepay"]):
    r   = SEC2 + 2 + i
    alt = i % 2 == 1
    ws.row_dimensions[r].height = 18
    row_label(ws, r, 2, status, alt)
    cnt  = f"COUNTIF({rng('S')},\"{status}\")"
    data_cell(ws, r, 3, cnt,                                                "#,##0",     alt)
    data_cell(ws, r, 4, f"{cnt}/{total_emi_expr}*100",                      "0.00\"%\"", alt)
    data_cell(ws, r, 5, f"SUMIF({rng('S')},\"{status}\",{rng('U')})",       "#,##0",     alt)
    data_cell(ws, r, 6, f"AVERAGEIF({rng('S')},\"{status}\",{rng('F')})",   "0.0",       alt)
    data_cell(ws, r, 7, f"SUMIF({rng('S')},\"{status}\",{rng('L')})",       "#,##0",     alt)

# =============================================================================
# SECTION 3 — DPD BUCKET ANALYSIS
# =============================================================================
SEC3 = SEC2 + 5 + 3    # 5 status rows + gap
section_header(ws, SEC3, 2, "3. DPD BUCKET ANALYSIS", span=6, bg=MID_BLUE)
ws.row_dimensions[SEC3].height = 22

for c, h in enumerate(["DPD Bucket", "EMI Count", "% of Total",
                        "Avg DPD", "Total Outstanding (Rs)", "Penal Interest (Rs)"], 2):
    col_header(ws, SEC3 + 1, c, h)
ws.row_dimensions[SEC3 + 1].height = 30

for i, bucket in enumerate(["Current", "DPD 1-30", "DPD 31-60", "DPD 61-90", "DPD 90+"]):
    r   = SEC3 + 2 + i
    alt = i % 2 == 1
    ws.row_dimensions[r].height = 18
    row_label(ws, r, 2, bucket, alt)
    cnt = f"COUNTIF({rng('G')},\"{bucket}\")"
    data_cell(ws, r, 3, cnt,                                                "#,##0",     alt)
    data_cell(ws, r, 4, f"{cnt}/{total_emi_expr}*100",                      "0.00\"%\"", alt)
    data_cell(ws, r, 5, f"AVERAGEIF({rng('G')},\"{bucket}\",{rng('F')})",   "0.0",       alt)
    data_cell(ws, r, 6, f"SUMIF({rng('G')},\"{bucket}\",{rng('H')})",       "#,##0",     alt)
    data_cell(ws, r, 7, f"SUMIF({rng('G')},\"{bucket}\",{rng('L')})",       "#,##0",     alt)

# =============================================================================
# SECTION 4 — LOAN STATUS SUMMARY
# =============================================================================
SEC4 = SEC3 + 5 + 3
section_header(ws, SEC4, 2, "4. LOAN STATUS SUMMARY", span=6, bg=MID_BLUE)
ws.row_dimensions[SEC4].height = 22

for c, h in enumerate(["Loan Status", "EMI Count", "% of Total",
                        "Total Outstanding (Rs)", "Total Paid (Rs)", "Default Penalty (Rs)"], 2):
    col_header(ws, SEC4 + 1, c, h)
ws.row_dimensions[SEC4 + 1].height = 30

for i, ls in enumerate(["STANDARD", "NPA"]):
    r   = SEC4 + 2 + i
    alt = i % 2 == 1
    ws.row_dimensions[r].height = 18
    row_label(ws, r, 2, ls, alt)
    cnt = f"COUNTIF({rng('Q')},\"{ls}\")"
    data_cell(ws, r, 3, cnt,                                            "#,##0",     alt)
    data_cell(ws, r, 4, f"{cnt}/{total_emi_expr}*100",                  "0.00\"%\"", alt)
    data_cell(ws, r, 5, f"SUMIF({rng('Q')},\"{ls}\",{rng('H')})",      "#,##0",     alt)
    data_cell(ws, r, 6, f"SUMIF({rng('Q')},\"{ls}\",{rng('U')})",      "#,##0",     alt)
    data_cell(ws, r, 7, f"SUMIF({rng('Q')},\"{ls}\",{rng('P')})",      "#,##0",     alt)

# =============================================================================
# SECTION 5 — PREPAYMENT ANALYSIS
# =============================================================================
SEC5 = SEC4 + 2 + 3
section_header(ws, SEC5, 2, "5. PREPAYMENT ANALYSIS", span=6, bg=MID_BLUE)
ws.row_dimensions[SEC5].height = 22

for c, h in enumerate(["Prepayment Type", "EMI Count", "% of Total",
                        "Total Prepaid (Rs)", "Total Charges (Rs)", "Avg Prepaid (Rs)"], 2):
    col_header(ws, SEC5 + 1, c, h)
ws.row_dimensions[SEC5 + 1].height = 30

for i, pt in enumerate(["Partial", "Full"]):
    r   = SEC5 + 2 + i
    alt = i % 2 == 1
    ws.row_dimensions[r].height = 18
    row_label(ws, r, 2, pt, alt)
    cnt     = f"COUNTIF({rng('O')},\"{pt}\")"
    sum_m   = f"SUMIF({rng('O')},\"{pt}\",{rng('M')})"
    data_cell(ws, r, 3, cnt,                                            "#,##0",     alt)
    data_cell(ws, r, 4, f"{cnt}/{total_emi_expr}*100",                  "0.00\"%\"", alt)
    data_cell(ws, r, 5, sum_m,                                          "#,##0",     alt)
    data_cell(ws, r, 6, f"SUMIF({rng('O')},\"{pt}\",{rng('N')})",      "#,##0",     alt)
    data_cell(ws, r, 7, f"IF({cnt}>0,{sum_m}/{cnt},0)",                 "#,##0",     alt)

# =============================================================================
# SECTION 6 — EMI NUMBER TREND (EMI 1-12)
# =============================================================================
SEC6 = SEC5 + 2 + 3
section_header(ws, SEC6, 2, "6. EMI NUMBER TREND (EMI 1-12)", span=6, bg=MID_BLUE)
ws.row_dimensions[SEC6].height = 22

for c, h in enumerate(["EMI #", "EMI Count", "Total Collected (Rs)",
                        "Total Penal Int (Rs)", "Total Default Pen (Rs)", "On-Time %"], 2):
    col_header(ws, SEC6 + 1, c, h)
ws.row_dimensions[SEC6 + 1].height = 30

for i in range(1, 13):
    r   = SEC6 + 1 + i
    alt = i % 2 == 0
    ws.row_dimensions[r].height = 18
    row_label(ws, r, 2, str(i), alt)
    cnt      = f"COUNTIF({rng('C')},{i})"
    on_time  = f"COUNTIFS({rng('C')},{i},{rng('S')},\"On Time\")"
    data_cell(ws, r, 3, cnt,                                        "#,##0",     alt)
    data_cell(ws, r, 4, f"SUMIF({rng('C')},{i},{rng('U')})",        "#,##0",     alt)
    data_cell(ws, r, 5, f"SUMIF({rng('C')},{i},{rng('L')})",        "#,##0",     alt)
    data_cell(ws, r, 6, f"SUMIF({rng('C')},{i},{rng('P')})",        "#,##0",     alt)
    data_cell(ws, r, 7, f"IF({cnt}>0,{on_time}/{cnt}*100,0)",        "0.0\"%\"",  alt)

# ── FREEZE PANES & ZOOM ───────────────────────────────────────────────────────
ws.freeze_panes = "B4"
ws.sheet_view.zoomScale = 90

# ── SAVE ──────────────────────────────────────────────────────────────────────
wb.save(OUT_PATH)
print(f"Dashboard saved: {OUT_PATH}")
